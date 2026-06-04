import logging
import io
from datetime import timedelta

from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.core.permissions import HasEmpresaPermission
from .models import OrdenServicio, SeguimientoProveedor, HistorialEstadoOS
from .serializers import (
    OrdenServicioSerializer,
    OrdenServicioListSerializer,
    OrdenServicioResumenSerializer,
    SeguimientoProveedorSerializer,
    HistorialEstadoOSSerializer,
)

logger = logging.getLogger(__name__)


class OrdenServicioViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def _get_empresa(self):
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied('Usuario no tiene empresa asignada.')
        return self.request.user.empresa

    def get_serializer_class(self):
        if self.action == 'list':
            return OrdenServicioListSerializer
        if self.action == 'resumen':
            return OrdenServicioResumenSerializer
        return OrdenServicioSerializer

    def get_queryset(self):
        empresa = self._get_empresa()
        qs = (
            OrdenServicio.objects
            .filter(empresa=empresa)
            .select_related('cliente', 'proveedor_reparacion', 'guia_remision', 'cotizacion', 'creado_por')
            .prefetch_related('items', 'seguimientos', 'historial')
        )

        params = self.request.query_params

        estado = params.get('estado')
        if estado:
            qs = qs.filter(estado=estado)

        proveedor = params.get('proveedor')
        if proveedor:
            qs = qs.filter(proveedor_reparacion_id=proveedor)

        cliente = params.get('cliente')
        if cliente:
            qs = qs.filter(cliente_id=cliente)

        prioridad = params.get('prioridad')
        if prioridad:
            qs = qs.filter(prioridad=prioridad)

        en_riesgo = params.get('en_riesgo')
        if en_riesgo in ('true', '1'):
            hoy = timezone.now().date()
            qs = qs.filter(
                fecha_entrega_estimada__lte=hoy + timedelta(days=10),
                estado__in=[
                    'diagnostico_pendiente', 'diagnostico_completado',
                    'cotizacion_enviada', 'cotizacion_aprobada',
                    'en_reparacion', 'reparacion_completada',
                ],
            )

        fecha_ingreso_desde = params.get('fecha_ingreso_desde')
        if fecha_ingreso_desde:
            qs = qs.filter(fecha_ingreso__gte=fecha_ingreso_desde)

        fecha_ingreso_hasta = params.get('fecha_ingreso_hasta')
        if fecha_ingreso_hasta:
            qs = qs.filter(fecha_ingreso__lte=fecha_ingreso_hasta)

        ordering = params.get('ordering', '-created_at')
        allowed_orderings = {
            'fecha_ingreso': 'fecha_ingreso',
            '-fecha_ingreso': '-fecha_ingreso',
            'dias_restantes_sla': 'fecha_entrega_estimada',
            '-dias_restantes_sla': '-fecha_entrega_estimada',
            'estado': 'estado',
            '-estado': '-estado',
            '-created_at': '-created_at',
        }
        qs = qs.order_by(allowed_orderings.get(ordering, '-created_at'))

        return qs

    def perform_create(self, serializer):
        empresa = self._get_empresa()
        serializer.save(empresa=empresa, creado_por=self.request.user)

    @action(detail=True, methods=['post'], url_path='cambiar_estado')
    def cambiar_estado(self, request, pk=None):
        orden = self.get_object()
        nuevo_estado = request.data.get('estado')
        notas = request.data.get('notas', '')

        estados_validos = [c[0] for c in OrdenServicio.ESTADO_CHOICES]
        if nuevo_estado not in estados_validos:
            return Response(
                {'error': f'Estado inválido. Opciones: {estados_validos}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        estado_anterior = orden.estado
        orden.estado = nuevo_estado

        hoy = timezone.now().date()
        if nuevo_estado == 'diagnostico_completado' and not orden.fecha_diagnostico:
            orden.fecha_diagnostico = hoy
        elif nuevo_estado == 'cotizacion_enviada' and not orden.fecha_cotizacion_enviada:
            orden.fecha_cotizacion_enviada = hoy
        elif nuevo_estado == 'cotizacion_aprobada' and not orden.fecha_aprobacion:
            orden.fecha_aprobacion = hoy
        elif nuevo_estado == 'en_reparacion' and not orden.fecha_inicio_reparacion:
            orden.fecha_inicio_reparacion = hoy
        elif nuevo_estado == 'entregado' and not orden.fecha_entrega_real:
            orden.fecha_entrega_real = hoy

        orden.save()

        HistorialEstadoOS.objects.create(
            orden=orden,
            estado_anterior=estado_anterior,
            estado_nuevo=nuevo_estado,
            usuario=request.user,
            notas=notas,
        )

        serializer = OrdenServicioSerializer(orden, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='agregar_seguimiento')
    def agregar_seguimiento(self, request, pk=None):
        orden = self.get_object()
        serializer = SeguimientoProveedorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        seguimiento = serializer.save(orden=orden, registrado_por=request.user)
        return Response(
            SeguimientoProveedorSerializer(seguimiento).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['get'], url_path='resumen')
    def resumen(self, request):
        empresa = self._get_empresa()
        estados_activos = [
            'diagnostico_pendiente', 'diagnostico_completado',
            'cotizacion_enviada', 'cotizacion_aprobada',
            'en_reparacion', 'reparacion_completada',
        ]
        qs_activas = OrdenServicio.objects.filter(empresa=empresa, estado__in=estados_activos)

        hoy = timezone.now().date()
        limite_critico = hoy + timedelta(days=10)
        en_riesgo_qs = qs_activas.filter(fecha_entrega_estimada__lte=limite_critico)
        vencidas_qs = qs_activas.filter(fecha_entrega_estimada__lt=hoy)

        por_estado = {}
        for codigo, label in OrdenServicio.ESTADO_CHOICES:
            count = OrdenServicio.objects.filter(empresa=empresa, estado=codigo).count()
            if count:
                por_estado[label] = count

        entregadas = OrdenServicio.objects.filter(
            empresa=empresa,
            estado='entregado',
            fecha_entrega_real__isnull=False,
        )
        tiempo_promedio = None
        if entregadas.exists():
            tiempos = [
                (o.fecha_entrega_real - o.fecha_ingreso).days
                for o in entregadas
                if o.fecha_entrega_real and o.fecha_ingreso
            ]
            if tiempos:
                tiempo_promedio = round(sum(tiempos) / len(tiempos), 1)

        from apps.servicios.models import ItemOrdenServicio
        from django.db.models import F, ExpressionWrapper, DecimalField
        items_con_margen = ItemOrdenServicio.objects.filter(
            orden__empresa=empresa,
            precio_cliente__isnull=False,
            costo_proveedor__isnull=False,
            precio_cliente__gt=0,
        )
        margen_promedio_pct = None
        if items_con_margen.exists():
            margenes = []
            for item in items_con_margen:
                if item.precio_cliente and item.precio_cliente > 0:
                    pct = float(item.precio_cliente - item.costo_proveedor) / float(item.precio_cliente) * 100
                    margenes.append(pct)
            if margenes:
                margen_promedio_pct = round(sum(margenes) / len(margenes), 1)

        por_proveedor = []
        proveedores = (
            qs_activas.values('proveedor_reparacion__id', 'proveedor_reparacion__razon_social')
            .annotate(total=Count('id'))
            .order_by('-total')
        )
        for p in proveedores:
            prov_id = p['proveedor_reparacion__id']
            prov_nombre = p['proveedor_reparacion__razon_social'] or str(prov_id)
            qs_prov = qs_activas.filter(proveedor_reparacion_id=prov_id)
            en_riesgo_prov = qs_prov.filter(fecha_entrega_estimada__lte=limite_critico).count()

            entregadas_prov = OrdenServicio.objects.filter(
                empresa=empresa,
                proveedor_reparacion_id=prov_id,
                estado='entregado',
                fecha_entrega_real__isnull=False,
            )
            tiempos_prov = [
                (o.fecha_entrega_real - o.fecha_ingreso).days
                for o in entregadas_prov
                if o.fecha_entrega_real and o.fecha_ingreso
            ]
            t_prom = round(sum(tiempos_prov) / len(tiempos_prov), 1) if tiempos_prov else None

            por_proveedor.append({
                'proveedor_id': prov_id,
                'proveedor': prov_nombre,
                'total': p['total'],
                'en_riesgo': en_riesgo_prov,
                'tiempo_promedio': t_prom,
            })

        data = {
            'total_activas': qs_activas.count(),
            'en_riesgo': en_riesgo_qs.count(),
            'vencidas': vencidas_qs.count(),
            'por_estado': por_estado,
            'tiempo_promedio_dias': tiempo_promedio,
            'margen_promedio_pct': margen_promedio_pct,
            'por_proveedor': por_proveedor,
        }
        return Response(data)

    @action(detail=False, methods=['get'], url_path='exportar_excel')
    def exportar_excel(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        empresa = self._get_empresa()
        ordenes = self.get_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Órdenes de Servicio'

        headers = [
            'N° Orden', 'Cliente', 'Proveedor', 'Prioridad', 'Estado',
            'Fecha Ingreso', 'Fecha Entrega Est.', 'Fecha Entrega Real',
            'Días Transcurridos', 'Días Restantes SLA', 'Estado SLA',
            'N° Items',
        ]

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        COLORES_SLA = {
            'vencido': 'FF0000',
            'critico': 'FF6600',
            'alerta': 'FFCC00',
            'ok': '00CC44',
            'sin_fecha': 'CCCCCC',
        }

        for row, orden in enumerate(ordenes, 2):
            ws.cell(row=row, column=1, value=orden.numero)
            ws.cell(row=row, column=2, value=str(orden.cliente))
            ws.cell(row=row, column=3, value=str(orden.proveedor_reparacion))
            ws.cell(row=row, column=4, value=orden.get_prioridad_display())
            ws.cell(row=row, column=5, value=orden.get_estado_display())
            ws.cell(row=row, column=6, value=str(orden.fecha_ingreso))
            ws.cell(row=row, column=7, value=str(orden.fecha_entrega_estimada) if orden.fecha_entrega_estimada else '')
            ws.cell(row=row, column=8, value=str(orden.fecha_entrega_real) if orden.fecha_entrega_real else '')
            ws.cell(row=row, column=9, value=orden.dias_transcurridos)
            ws.cell(row=row, column=10, value=orden.dias_restantes_sla)
            ws.cell(row=row, column=11, value=orden.estado_sla)
            ws.cell(row=row, column=12, value=orden.items.count())

            color = COLORES_SLA.get(orden.estado_sla, 'FFFFFF')
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws.cell(row=row, column=11).fill = fill

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="ordenes_servicio_{empresa.ruc}_{timezone.now().date()}.xlsx"'
        )
        return response
