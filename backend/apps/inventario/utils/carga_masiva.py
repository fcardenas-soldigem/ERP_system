"""
Módulo de Carga Masiva de Inventario con Detección Automática de Tipo de Producto
================================================================================

Este módulo procesa archivos Excel para cargar productos al inventario,
detectando automáticamente si son materias primas o productos terminados
basándose en la categoría especificada.

CATEGORÍAS PARA MATERIA PRIMA:
- MP, Materia Prima, Materias Primas, M.P, M P
- Insumo, Insumos
- Material, Materiales
- Ingrediente, Componente

CATEGORÍAS PARA PRODUCTO TERMINADO:
- PT, Producto Terminado, Productos Terminados, P.T, P T
- Cualquier otra categoría (Electrónicos, Muebles, Ropa, etc.)
"""

import pandas as pd
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Dict, List, Any, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from django.db import transaction


class DetectorTipoProducto:
    """
    Clase para detectar automáticamente si un producto es Materia Prima
    o Producto Terminado basándose en la categoría
    """
    
    # Palabras clave que indican MATERIA PRIMA
    CATEGORIAS_MATERIA_PRIMA = [
        'mp', 'm.p', 'm p',
        'materia prima', 'materiaprima', 'materia_prima', 'materias primas',
        'mat prima', 'mat. prima',
        'insumo', 'insumos',
        'material', 'materiales',
        'raw', 'raw material', 'raw materials',
        'ingrediente', 'ingredientes',
        'componente', 'componentes',
    ]
    
    # Palabras clave que indican PRODUCTO TERMINADO
    CATEGORIAS_PRODUCTO_TERMINADO = [
        'pt', 'p.t', 'p t',
        'producto terminado', 'productoterminado', 'producto_terminado', 'productos terminados',
        'prod terminado', 'prod. terminado',
        'finished', 'finished product', 'finished products',
        'final',
    ]
    
    @classmethod
    def normalizar_categoria(cls, categoria: str) -> str:
        """Normaliza una categoría removiendo espacios, guiones y underscores"""
        if not categoria or pd.isna(categoria):
            return ''
        categoria = str(categoria).lower().strip()
        categoria = re.sub(r'[_\-\s]+', ' ', categoria).strip()
        return categoria
    
    @classmethod
    def detectar(cls, categoria: str) -> Dict[str, Any]:
        """
        Detecta el tipo de producto basándose en la categoría
        
        Returns:
            Dict con 'tipo', 'tipo_codigo', 'confianza' y 'metodo'
        """
        categoria_normalizada = cls.normalizar_categoria(categoria)
        
        # PASO 1: Verificar contra lista de materias primas (exacta)
        if categoria_normalizada in cls.CATEGORIAS_MATERIA_PRIMA:
            return {
                'tipo': 'Materia Prima',
                'tipo_codigo': 'RAW',
                'confianza': 0.99,
                'metodo': 'categoria_exacta',
                'categoria_original': categoria
            }
        
        # PASO 2: Verificar contra lista de productos terminados explícitos
        if categoria_normalizada in cls.CATEGORIAS_PRODUCTO_TERMINADO:
            return {
                'tipo': 'Producto Terminado',
                'tipo_codigo': 'FINISHED',
                'confianza': 0.99,
                'metodo': 'categoria_exacta',
                'categoria_original': categoria
            }
        
        # PASO 3: Búsqueda parcial de palabras clave para materia prima
        for palabra_clave in ['materia', 'insumo', 'material', 'ingrediente', 'componente', 'raw']:
            if palabra_clave in categoria_normalizada:
                return {
                    'tipo': 'Materia Prima',
                    'tipo_codigo': 'RAW',
                    'confianza': 0.85,
                    'metodo': 'palabra_clave_parcial',
                    'categoria_original': categoria
                }
        
        # PASO 4: Búsqueda parcial para producto terminado
        for palabra_clave in ['producto', 'terminado', 'final', 'finished']:
            if palabra_clave in categoria_normalizada:
                return {
                    'tipo': 'Producto Terminado',
                    'tipo_codigo': 'FINISHED',
                    'confianza': 0.85,
                    'metodo': 'palabra_clave_parcial',
                    'categoria_original': categoria
                }
        
        # PASO 5: Por defecto, cualquier otra categoría es producto terminado
        return {
            'tipo': 'Producto Terminado',
            'tipo_codigo': 'FINISHED',
            'confianza': 0.75,
            'metodo': 'default_categoria_comercial',
            'categoria_original': categoria,
            'nota': 'Categoría comercial detectada como producto terminado'
        }


# Función simple para compatibilidad con código existente
def detectar_tipo_producto(categoria_nombre):
    """
    Detecta automáticamente el tipo de producto basándose en la categoría.
    
    Args:
        categoria_nombre: Nombre de la categoría del producto
        
    Returns:
        str: 'RAW' para materias primas, 'FINISHED' para productos terminados
    """
    resultado = DetectorTipoProducto.detectar(categoria_nombre)
    return resultado['tipo_codigo']


def detectar_tipo_producto_detallado(categoria_nombre):
    """
    Versión detallada que retorna información completa de la detección.
    
    Returns:
        Dict con tipo, confianza, método, etc.
    """
    return DetectorTipoProducto.detectar(categoria_nombre)


def validar_fila(fila, numero_fila, tipo_producto):
    """
    Valida una fila del Excel según el tipo de producto detectado.
    
    Args:
        fila: Diccionario con los datos de la fila
        numero_fila: Número de fila para mensajes de error
        tipo_producto: 'RAW' o 'FINISHED'
        
    Returns:
        tuple: (es_valido, errores, warnings)
    """
    errores = []
    warnings = []
    
    # Campos obligatorios para todos
    if not fila.get('sku') or pd.isna(fila.get('sku')):
        errores.append(f"Fila {numero_fila}: SKU es obligatorio")
    
    if not fila.get('nombre') or pd.isna(fila.get('nombre')):
        errores.append(f"Fila {numero_fila}: Nombre es obligatorio")
    
    if not fila.get('categoria') or pd.isna(fila.get('categoria')):
        errores.append(f"Fila {numero_fila}: Categoría es obligatoria")
    
    # Precio de compra siempre obligatorio
    precio_compra = fila.get('precio_compra')
    if precio_compra is None or pd.isna(precio_compra):
        errores.append(f"Fila {numero_fila}: Precio de compra es obligatorio")
    else:
        try:
            pc = Decimal(str(precio_compra))
            if pc <= 0:
                errores.append(f"Fila {numero_fila}: Precio de compra debe ser mayor a 0")
        except (InvalidOperation, ValueError):
            errores.append(f"Fila {numero_fila}: Precio de compra inválido")
    
    # Validaciones específicas por tipo
    if tipo_producto == 'RAW':
        # Materia prima: precio_venta se ignora
        precio_venta = fila.get('precio_venta')
        if precio_venta and not pd.isna(precio_venta) and float(precio_venta) > 0:
            warnings.append(
                f"Fila {numero_fila}: Precio de venta ignorado para materia prima '{fila.get('nombre')}'"
            )
    else:
        # Producto terminado: precio_venta obligatorio
        precio_venta = fila.get('precio_venta')
        if precio_venta is None or pd.isna(precio_venta):
            errores.append(f"Fila {numero_fila}: Precio de venta es obligatorio para productos terminados")
        else:
            try:
                pv = Decimal(str(precio_venta))
                if pv <= 0:
                    errores.append(f"Fila {numero_fila}: Precio de venta debe ser mayor a 0")
                # Validar que precio_venta >= precio_compra
                if precio_compra and not pd.isna(precio_compra):
                    pc = Decimal(str(precio_compra))
                    if pv < pc:
                        warnings.append(
                            f"Fila {numero_fila}: Precio de venta ({pv}) es menor al precio de compra ({pc})"
                        )
            except (InvalidOperation, ValueError):
                errores.append(f"Fila {numero_fila}: Precio de venta inválido")
    
    return len(errores) == 0, errores, warnings


def procesar_carga_masiva_excel(archivo_path, empresa):
    """
    Procesa un archivo Excel de carga masiva de inventario.
    
    Args:
        archivo_path: Ruta al archivo Excel o objeto BytesIO
        empresa: Instancia de Empresa para asociar los productos
        
    Returns:
        dict: Resultado del procesamiento con resumen, errores y productos creados
    """
    from apps.inventario.models import (
        Producto, Categoria, Almacen, Stock,
        InventarioMateriasPrimas, InventarioProductosTerminados
    )
    
    resultado = {
        'success': False,
        'resumen': {
            'filas_procesadas': 0,
            'productos_validos': 0,
            'materias_primas': 0,
            'productos_terminados': 0,
            'errores': 0,
            'warnings': 0,
            'skus_duplicados_archivo': 0,
            'skus_existentes_bd': 0
        },
        'errores': [],
        'warnings': [],
        'productos_creados': [],
        'detalle_deteccion': []  # Info de confianza de detección
    }
    
    try:
        # Leer el Excel
        df = pd.read_excel(archivo_path, sheet_name='Template Productos')
        
        # Normalizar nombres de columnas
        df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')
        
        # Mapeo de columnas alternativas
        column_mapping = {
            'código': 'sku',
            'codigo': 'sku',
            'nombre_producto': 'nombre',
            'descripción': 'descripcion',
            'categoría': 'categoria',
            'almacén': 'almacen',
            'precio_de_compra': 'precio_compra',
            'precio_de_venta': 'precio_venta',
            'stock_mínimo': 'stock_minimo',
            'stock_máximo': 'stock_maximo',
            'unidad_de_medida': 'unidad_medida'
        }
        df.rename(columns=column_mapping, inplace=True)
        
        # Filtrar filas vacías
        df = df.dropna(subset=['sku'], how='all')
        df = df[df['sku'].astype(str).str.strip() != '']
        
        resultado['resumen']['filas_procesadas'] = len(df)
        
        # ========================================
        # DETECTAR SKUs DUPLICADOS EN EL ARCHIVO
        # ========================================
        skus_en_archivo = df['sku'].astype(str).str.strip().tolist()
        skus_duplicados = [sku for sku in set(skus_en_archivo) if skus_en_archivo.count(sku) > 1]
        
        if skus_duplicados:
            resultado['resumen']['skus_duplicados_archivo'] = len(skus_duplicados)
            for sku in skus_duplicados:
                resultado['errores'].append(
                    f"SKU '{sku}' está duplicado en el archivo Excel"
                )
        
        # ========================================
        # DETECTAR SKUs QUE YA EXISTEN EN LA BD
        # ========================================
        skus_existentes = set(
            Producto.objects.filter(
                empresa=empresa,
                sku__in=skus_en_archivo
            ).values_list('sku', flat=True)
        )
        
        if skus_existentes:
            resultado['resumen']['skus_existentes_bd'] = len(skus_existentes)
            for sku in skus_existentes:
                resultado['warnings'].append(
                    f"SKU '{sku}' ya existe en la base de datos, será omitido"
                )
        
        productos_a_crear = []
        
        # Procesar cada fila
        for idx, row in df.iterrows():
            numero_fila = idx + 2  # +2 porque Excel empieza en 1 y tiene header
            
            # Convertir fila a diccionario
            fila = row.to_dict()
            sku_actual = str(fila.get('sku', '')).strip()
            
            # Saltar filas vacías
            if pd.isna(fila.get('sku')) and pd.isna(fila.get('nombre')):
                continue
            
            # Saltar SKUs duplicados en archivo
            if sku_actual in skus_duplicados:
                continue
            
            # Saltar SKUs que ya existen en BD
            if sku_actual in skus_existentes:
                continue
            
            # Detectar tipo de producto (versión detallada)
            tipo_info = detectar_tipo_producto_detallado(fila.get('categoria'))
            tipo_producto = tipo_info['tipo_codigo']
            
            # Guardar detalle de detección
            resultado['detalle_deteccion'].append({
                'fila': numero_fila,
                'sku': sku_actual,
                'categoria': fila.get('categoria'),
                'tipo_detectado': tipo_info['tipo'],
                'confianza': tipo_info['confianza'],
                'metodo': tipo_info['metodo']
            })
            
            # Warning si confianza es baja
            if tipo_info['confianza'] < 0.80:
                resultado['warnings'].append(
                    f"Fila {numero_fila}: Clasificación de '{sku_actual}' con confianza {tipo_info['confianza']:.0%}. Revisar categoría."
                )
            
            # Validar fila
            es_valido, errores, warnings = validar_fila(fila, numero_fila, tipo_producto)
            
            resultado['errores'].extend(errores)
            resultado['warnings'].extend(warnings)
            
            if es_valido:
                productos_a_crear.append({
                    'fila': numero_fila,
                    'datos': fila,
                    'tipo_producto': tipo_producto,
                    'tipo_info': tipo_info
                })
            else:
                resultado['resumen']['errores'] += len(errores)
        
        resultado['resumen']['warnings'] = len(resultado['warnings'])
        resultado['resumen']['errores'] = len(resultado['errores'])
        
        # Si hay errores críticos, no procesar
        if resultado['resumen']['errores'] > 0:
            resultado['success'] = False
            return resultado
        
        # Crear productos en una transacción
        with transaction.atomic():
            for item in productos_a_crear:
                fila = item['datos']
                tipo = item['tipo_producto']
                tipo_info = item['tipo_info']
                
                # Obtener o crear categoría
                categoria_nombre = str(fila.get('categoria', '')).strip()
                categoria, _ = Categoria.objects.get_or_create(
                    empresa=empresa,
                    nombre=categoria_nombre,
                    defaults={'descripcion': f'Categoría {categoria_nombre}'}
                )
                
                # Obtener almacén (si se especifica)
                almacen = None
                almacen_nombre = fila.get('almacen')
                if almacen_nombre and not pd.isna(almacen_nombre):
                    almacen, _ = Almacen.objects.get_or_create(
                        empresa=empresa,
                        nombre=str(almacen_nombre).strip(),
                        defaults={'direccion': ''}
                    )
                else:
                    # Usar almacén por defecto
                    almacen = Almacen.objects.filter(empresa=empresa).first()
                
                # Preparar datos del producto
                sku = str(fila.get('sku')).strip()
                precio_compra = Decimal(str(fila.get('precio_compra', 0)))
                
                # Precio de venta según tipo
                if tipo == 'RAW':
                    precio_venta = Decimal('0')
                else:
                    precio_venta = Decimal(str(fila.get('precio_venta', 0)))
                
                # Crear producto
                producto = Producto.objects.create(
                    empresa=empresa,
                    sku=sku,
                    nombre=str(fila.get('nombre', '')).strip(),
                    descripcion=str(fila.get('descripcion', '')).strip() if not pd.isna(fila.get('descripcion')) else '',
                    tipo_producto=tipo,
                    categoria=categoria,
                    almacen=almacen,
                    precio_compra=precio_compra,
                    precio_venta=precio_venta,
                    moneda=str(fila.get('moneda', 'PEN')).upper() if not pd.isna(fila.get('moneda')) else 'PEN',
                    stock_minimo=Decimal(str(fila.get('stock_minimo', 0))) if not pd.isna(fila.get('stock_minimo')) else 0,
                    stock_maximo=Decimal(str(fila.get('stock_maximo', 0))) if not pd.isna(fila.get('stock_maximo')) else 0,
                    unidad_medida=str(fila.get('unidad_medida', 'unidad')).lower() if not pd.isna(fila.get('unidad_medida')) else 'unidad',
                )
                
                # Stock inicial
                stock_inicial = Decimal(str(fila.get('stock_inicial', 0))) if not pd.isna(fila.get('stock_inicial')) else Decimal('0')
                
                # Crear en inventario correspondiente
                if tipo == 'RAW' and almacen:
                    InventarioMateriasPrimas.objects.create(
                        empresa=empresa,
                        producto=producto,
                        almacen=almacen,
                        cantidad_disponible=stock_inicial,
                        cantidad_reservada=0,
                        costo_unitario_promedio=precio_compra,
                        stock_minimo=producto.stock_minimo,
                        stock_maximo=producto.stock_maximo
                    )
                    resultado['resumen']['materias_primas'] += 1
                    
                elif tipo == 'FINISHED' and almacen:
                    InventarioProductosTerminados.objects.create(
                        empresa=empresa,
                        producto=producto,
                        almacen=almacen,
                        cantidad_disponible=stock_inicial,
                        cantidad_reservada=0,
                        costo_produccion_unitario=precio_compra,
                        precio_venta_sugerido=precio_venta,
                        stock_minimo=producto.stock_minimo,
                        stock_maximo=producto.stock_maximo
                    )
                    resultado['resumen']['productos_terminados'] += 1
                
                # También crear en Stock para compatibilidad
                if almacen:
                    Stock.objects.create(
                        empresa=empresa,
                        producto=producto,
                        almacen=almacen,
                        cantidad=stock_inicial
                    )
                
                resultado['productos_creados'].append({
                    'sku': sku,
                    'nombre': producto.nombre,
                    'tipo': tipo,
                    'tipo_display': tipo_info['tipo'],
                    'categoria': categoria_nombre,
                    'confianza_deteccion': tipo_info['confianza']
                })
                resultado['resumen']['productos_validos'] += 1
        
        resultado['success'] = True
        
    except Exception as e:
        resultado['success'] = False
        resultado['errores'].append(f"Error al procesar archivo: {str(e)}")
    
    return resultado


def generar_plantilla_excel():
    """
    Genera una plantilla Excel mejorada para carga masiva de inventario.
    
    Returns:
        BytesIO: Archivo Excel en memoria
    """
    wb = Workbook()
    
    # ========================================
    # HOJA 1: INSTRUCCIONES
    # ========================================
    ws_instrucciones = wb.active
    ws_instrucciones.title = "📋 INSTRUCCIONES"
    
    # Estilos
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    
    instrucciones = [
        ["GUÍA DE CARGA MASIVA DE INVENTARIO"],
        [""],
        ["DETECCIÓN AUTOMÁTICA DE TIPO DE PRODUCTO"],
        ["El sistema detecta automáticamente si un producto es Materia Prima o Producto Terminado"],
        ["basándose en la categoría que especifiques."],
        [""],
        ["CATEGORÍAS PARA MATERIA PRIMA:"],
        ["  • MP, Materia Prima, Insumo, Insumos, Material, Materiales"],
        ["  → precio_venta será IGNORADO (se pondrá en 0)"],
        [""],
        ["CATEGORÍAS PARA PRODUCTO TERMINADO:"],
        ["  • PT, Producto Terminado, o cualquier otra categoría"],
        ["  → precio_venta es OBLIGATORIO"],
        [""],
        ["CAMPOS OBLIGATORIOS:"],
        ["  • sku - Código único del producto"],
        ["  • nombre - Nombre del producto"],
        ["  • categoria - Determina el tipo de producto"],
        ["  • precio_compra - Siempre obligatorio"],
        ["  • precio_venta - Solo para productos terminados"],
        [""],
        ["CAMPOS OPCIONALES:"],
        ["  • descripcion, almacen, moneda (PEN/USD)"],
        ["  • stock_minimo, stock_maximo, stock_inicial"],
        ["  • unidad_medida (unidad, kilo, litro, metro, etc.)"],
    ]
    
    for i, row in enumerate(instrucciones):
        ws_instrucciones.append(row)
        if i == 0:
            ws_instrucciones.cell(row=1, column=1).font = header_font
            ws_instrucciones.cell(row=1, column=1).fill = header_fill
        elif row and row[0].startswith("CATEGORÍAS") or row[0].startswith("CAMPOS") or row[0].startswith("DETECCIÓN"):
            ws_instrucciones.cell(row=i+1, column=1).font = subheader_font
    
    ws_instrucciones.column_dimensions['A'].width = 80
    
    # ========================================
    # HOJA 2: TEMPLATE PRODUCTOS
    # ========================================
    ws_template = wb.create_sheet("Template Productos")
    
    headers = [
        'sku', 'nombre', 'descripcion', 'categoria', 'almacen',
        'precio_compra', 'precio_venta', 'moneda', 'stock_minimo',
        'stock_maximo', 'stock_inicial', 'unidad_medida'
    ]
    
    # Estilo de headers
    header_fill_green = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws_template.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill_green
        cell.alignment = Alignment(horizontal='center')
    
    # Ejemplos de datos
    ejemplos = [
        # Materias primas
        ['MP-ACERO-001', 'Acero Inoxidable 304', 'Lámina de acero para producción', 'MP', 'Almacén Principal', 45.50, 0, 'PEN', 100, 500, 200, 'kilo'],
        ['MP-MADERA-001', 'Madera Roble', 'Tablones de roble 2x4', 'Materia Prima', 'Almacén Principal', 120.00, 0, 'PEN', 50, 200, 100, 'unidad'],
        ['INS-TORN-001', 'Tornillos Acero', 'Tornillos 3/4 pulgada', 'Insumo', 'Almacén Principal', 0.15, 0, 'PEN', 1000, 5000, 2000, 'unidad'],
        # Productos terminados
        ['PT-MESA-001', 'Mesa de Comedor', 'Mesa de madera para 6 personas', 'Muebles', 'Almacén Principal', 450.00, 750.00, 'PEN', 5, 20, 10, 'unidad'],
        ['PT-SILLA-001', 'Silla Ejecutiva', 'Silla ergonómica de oficina', 'Muebles', 'Almacén Principal', 180.00, 350.00, 'PEN', 10, 50, 25, 'unidad'],
        ['ELEC-LAPTOP-001', 'Laptop Empresarial', 'Laptop 15" Core i7', 'Electrónicos', 'Almacén Principal', 2500.00, 3500.00, 'USD', 5, 20, 8, 'unidad'],
    ]
    
    for row_data in ejemplos:
        ws_template.append(row_data)
    
    # Ajustar anchos de columna
    column_widths = [15, 25, 35, 15, 20, 15, 15, 10, 12, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws_template.column_dimensions[chr(64 + i)].width = width
    
    # ========================================
    # HOJA 3: CATEGORÍAS VÁLIDAS
    # ========================================
    ws_categorias = wb.create_sheet("Categorías Válidas")
    
    ws_categorias.append(['TIPO', 'CATEGORÍAS VÁLIDAS', 'RESULTADO'])
    ws_categorias.cell(row=1, column=1).font = header_font_white
    ws_categorias.cell(row=1, column=1).fill = header_fill_green
    ws_categorias.cell(row=1, column=2).font = header_font_white
    ws_categorias.cell(row=1, column=2).fill = header_fill_green
    ws_categorias.cell(row=1, column=3).font = header_font_white
    ws_categorias.cell(row=1, column=3).fill = header_fill_green
    
    categorias_info = [
        ['MATERIA PRIMA', 'MP, Materia Prima, Insumo, Material', 'precio_venta = 0 (ignorado)'],
        ['MATERIA PRIMA', 'Ingrediente, Componente, Raw Material', 'precio_venta = 0 (ignorado)'],
        ['PRODUCTO TERMINADO', 'PT, Producto Terminado, Finished', 'precio_venta OBLIGATORIO'],
        ['PRODUCTO TERMINADO', 'Muebles, Electrónicos, Ropa, etc.', 'precio_venta OBLIGATORIO'],
        ['PRODUCTO TERMINADO', 'Cualquier otra categoría', 'precio_venta OBLIGATORIO'],
    ]
    
    for row in categorias_info:
        ws_categorias.append(row)
    
    ws_categorias.column_dimensions['A'].width = 20
    ws_categorias.column_dimensions['B'].width = 40
    ws_categorias.column_dimensions['C'].width = 30
    
    # ========================================
    # HOJA 4: EJEMPLOS DETALLADOS
    # ========================================
    ws_ejemplos = wb.create_sheet("📚 Ejemplos Detallados")
    
    ejemplos_detallados = [
        ["EJEMPLO 1: MATERIA PRIMA"],
        ["Campo", "Valor", "Explicación"],
        ["sku", "MP-ACERO-001", "Código único, prefijo MP recomendado"],
        ["nombre", "Acero Inoxidable", "Nombre descriptivo"],
        ["categoria", "MP", "← Esto lo marca como MATERIA PRIMA"],
        ["precio_compra", "45.50", "Costo de adquisición"],
        ["precio_venta", "0 o vacío", "Se ignora automáticamente"],
        [""],
        ["EJEMPLO 2: PRODUCTO TERMINADO"],
        ["Campo", "Valor", "Explicación"],
        ["sku", "PT-MESA-001", "Código único, prefijo PT recomendado"],
        ["nombre", "Mesa de Comedor", "Nombre del producto final"],
        ["categoria", "Muebles", "← Cualquier categoría = Producto Terminado"],
        ["precio_compra", "450.00", "Costo de producción/compra"],
        ["precio_venta", "750.00", "OBLIGATORIO, precio al cliente"],
    ]
    
    for row in ejemplos_detallados:
        ws_ejemplos.append(row)
    
    ws_ejemplos.column_dimensions['A'].width = 20
    ws_ejemplos.column_dimensions['B'].width = 25
    ws_ejemplos.column_dimensions['C'].width = 45
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
