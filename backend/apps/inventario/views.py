from rest_framework import generics, permissions, viewsets, views, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models import Sum, Count, Q, Case, When, ExpressionWrapper, DecimalField, F, Prefetch
from django.db import transaction
from django.utils import timezone
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.core.cache import cache
from decimal import Decimal
import openpyxl
from datetime import datetime, timedelta
import json
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.shortcuts import get_object_or_404

# Cache utilities
from apps.core.cache_utils import cache_manager, make_cache_key, get_cache_timeout

from .models import (
    Almacen, 
    Categoria, 
    Producto, 
    Stock, 
    AjusteInventario, 
    MovimientoInventario,
    InventarioMateriasPrimas,
    InventarioProductosTerminados
)
from .serializers import (
    AlmacenSerializer, 
    CategoriaSerializer, 
    ProductoSerializer, 
    StockSerializer,
    AjusteInventarioSerializer,
    MovimientoInventarioSerializer,
    KardexSerializer,
    KardexResumenSerializer,
    MovimientoInventarioCreateSerializer,
    InventarioMateriasPrimasSerializer,
    InventarioMateriasPrimasCreateSerializer,
    InventarioProductosTerminadosSerializer,
    InventarioProductosTerminadosCreateSerializer,
    EntradaCompraSerializer,
    SalidaProduccionSerializer,
    EntradaProduccionSerializer,
    SalidaVentaSerializer,
    AlertaStockSerializer,
    ResumenInventarioSeparadoSerializer
)
from apps.core.permissions import HasEmpresaPermission

class AlmacenViewSet(viewsets.ModelViewSet):
    serializer_class = AlmacenSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        return Almacen.objects.filter(empresa=self.request.user.empresa).order_by('id')

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

class CategoriaViewSet(viewsets.ModelViewSet):
    serializer_class = CategoriaSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        return Categoria.objects.filter(empresa=self.request.user.empresa)

    def list(self, request, *args, **kwargs):
        empresa_id = request.user.empresa_id
        cache_key = f'categorias_list_{empresa_id}'
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)
        response = super().list(request, *args, **kwargs)
        cache.set(cache_key, response.data, get_cache_timeout('categorias'))
        return response

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)
        cache.delete(f'categorias_list_{self.request.user.empresa_id}')

    def perform_update(self, serializer):
        serializer.save()
        cache.delete(f'categorias_list_{self.request.user.empresa_id}')

    def perform_destroy(self, instance):
        instance.delete()
        cache.delete(f'categorias_list_{self.request.user.empresa_id}')

class ProductoViewSet(viewsets.ModelViewSet):
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['categoria', 'almacen', 'tipo_producto']
    search_fields = ['nombre', 'sku', 'descripcion']
    ordering_fields = ['nombre', 'sku', 'stock_total', 'precio_venta']
    ordering = ['sku']

    def get_queryset(self):
        # Optimizar con select_related y prefetch_related
        queryset = Producto.objects.filter(
            empresa=self.request.user.empresa, 
            is_active=True
        ).select_related(
            'categoria', 
            'almacen'
        ).prefetch_related(
            Prefetch('stocks', queryset=Stock.objects.select_related('almacen')),
        )
        return queryset

    def list(self, request, *args, **kwargs):
        """Override list para agregar caching"""
        empresa_id = request.user.empresa_id
        cache_key = make_cache_key('productos_list', empresa_id)
        
        # Intentar obtener del cache
        cached_data = cache.get(cache_key)
        if cached_data is not None and not request.query_params:
            return Response(cached_data)
        
        # Si no hay cache o hay filtros, ejecutar query normal
        response = super().list(request, *args, **kwargs)
        
        # Solo cachear si no hay filtros
        if not request.query_params:
            cache.set(cache_key, response.data, get_cache_timeout('productos'))
        
        return response

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)
        # Invalidar cache
        cache_manager.invalidate_productos(self.request.user.empresa_id)

    def perform_update(self, serializer):
        serializer.save()
        # Invalidar cache
        cache_manager.invalidate_productos(self.request.user.empresa_id)

    def perform_destroy(self, instance):
        instance.delete()
        # Invalidar cache
        cache_manager.invalidate_productos(self.request.user.empresa_id)

    @action(detail=False, methods=['get'])
    def stock_bajo(self, request):
        productos = self.get_queryset().filter(
            stock_total__lte=F('stock_minimo')
        )
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def materias_primas(self, request):
        """Obtiene solo productos que pertenecen a la categoría 'Materia Prima' o 'Insumo'"""
        from django.db.models import Q
        productos = self.get_queryset().filter(
            Q(categoria__nombre__icontains='materia prima') |
            Q(categoria__nombre__icontains='insumo')
        )
        serializer = self.get_serializer(productos, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def check_sku(self, request):
        """Verifica si un SKU ya existe"""
        sku = request.query_params.get('sku', '').strip()
        if not sku:
            return Response({'exists': False})
        
        exists = Producto.objects.filter(
            empresa=request.user.empresa,
            sku__iexact=sku
        ).exists()
        
        return Response({'exists': exists})
    
    @action(detail=False, methods=['get'])
    def valor_total_adquisicion(self, request):
        """Obtiene el valor total de adquisición del inventario"""
        empresa = request.user.empresa
        
        # Calcular valor total de adquisición (precio_compra * stock_total)
        productos = Producto.objects.filter(empresa=empresa, is_active=True)
        valor_total = sum(
            float(producto.precio_compra or 0) * float(producto.stock_total or 0)
            for producto in productos
        )
        
        return Response({'valor_total_adquisicion': valor_total})

class StockViewSet(viewsets.ModelViewSet):
    serializer_class = StockSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        return Stock.objects.filter(empresa=self.request.user.empresa).order_by('id')

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

class AjusteInventarioViewSet(viewsets.ModelViewSet):
    serializer_class = AjusteInventarioSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        return AjusteInventario.objects.filter(empresa=self.request.user.empresa).order_by('id')

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

class ResumenInventario(views.APIView):
    """Vista para obtener resumen general del inventario"""
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    
    def get(self, request):
        empresa = request.user.empresa
        
        # Estadísticas generales
        total_productos = Producto.objects.filter(empresa=empresa, is_active=True).count()
        total_categorias = Categoria.objects.filter(empresa=empresa).count()
        total_almacenes = Almacen.objects.filter(empresa=empresa).count()
        
        # Productos con stock bajo
        productos_stock_bajo = Producto.objects.filter(
            empresa=empresa,
            is_active=True,
            stock_total__lte=F('stock_minimo')
        ).count()
        
        # Productos sin stock
        productos_sin_stock = Producto.objects.filter(
            empresa=empresa,
            is_active=True,
            stock_total=0
        ).count()
        
        # Valor total del inventario
        valor_inventario = Producto.objects.filter(
            empresa=empresa,
            is_active=True
        ).aggregate(
            valor_compra=Sum(F('stock_total') * F('precio_compra')),
            valor_venta=Sum(F('stock_total') * F('precio_venta'))
        )
        
        # Top productos por valor
        top_productos = Producto.objects.filter(
            empresa=empresa,
            is_active=True,
            stock_total__gt=0
        ).annotate(
            valor_stock=ExpressionWrapper(
                F('stock_total') * F('precio_venta'),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        ).order_by('-valor_stock')[:5]
        
        # Movimientos recientes (últimos 7 días)
        fecha_limite = timezone.now() - timedelta(days=7)
        movimientos_recientes = MovimientoInventario.objects.filter(
            empresa=empresa,
            fecha__gte=fecha_limite
        ).count()
        
        data = {
            'estadisticas_generales': {
                'total_productos': total_productos,
                'total_categorias': total_categorias,
                'total_almacenes': total_almacenes,
                'productos_stock_bajo': productos_stock_bajo,
                'productos_sin_stock': productos_sin_stock,
                'movimientos_recientes': movimientos_recientes
            },
            'valor_inventario': {
                'valor_compra': valor_inventario['valor_compra'] or 0,
                'valor_venta': valor_inventario['valor_venta'] or 0
            },
            'top_productos': ProductoSerializer(top_productos, many=True).data
        }
        
        return Response(data)

class ProductoEstadisticasView(views.APIView):
    """Vista para estadísticas específicas de productos"""
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    
    def get(self, request):
        empresa = request.user.empresa
        
        # Distribución por categorías
        categorias_stats = Categoria.objects.filter(empresa=empresa).annotate(
            total_productos=Count('productos', filter=Q(productos__is_active=True)),
            valor_total=Sum(
                ExpressionWrapper(
                    F('productos__stock_total') * F('productos__precio_venta'),
                    output_field=DecimalField(max_digits=12, decimal_places=2)
                ),
                filter=Q(productos__is_active=True)
            )
        ).order_by('-total_productos')
        
        # Productos más vendidos (basado en movimientos de salida)
        productos_vendidos = MovimientoInventario.objects.filter(
            empresa=empresa,
            tipo_movimiento='salida',
            tipo_documento='venta',
            fecha__gte=timezone.now() - timedelta(days=30)
        ).values('producto__nombre', 'producto__sku').annotate(
            total_vendido=Sum('cantidad_salida')
        ).order_by('-total_vendido')[:10]
        
        # Rotación de inventario por producto
        rotacion_productos = []
        for producto in Producto.objects.filter(empresa=empresa, is_active=True)[:20]:
            ventas_30d = MovimientoInventario.objects.filter(
                empresa=empresa,
                producto=producto,
                tipo_movimiento='salida',
                tipo_documento='venta',
                fecha__gte=timezone.now() - timedelta(days=30)
            ).aggregate(total=Sum('cantidad_salida'))['total'] or 0
            
            if producto.stock_total > 0 and ventas_30d > 0:
                rotacion = (ventas_30d / producto.stock_total) * 12  # Anualizado
                rotacion_productos.append({
                    'producto': producto.nombre,
                    'sku': producto.sku,
                    'rotacion_anual': round(float(rotacion), 2),
                    'stock_actual': float(producto.stock_total),
                    'ventas_30d': float(ventas_30d)
                })
        
        # Ordenar por rotación
        rotacion_productos.sort(key=lambda x: x['rotacion_anual'], reverse=True)
        
        data = {
            'categorias_stats': [
                {
                    'categoria': cat.nombre,
                    'total_productos': cat.total_productos,
                    'valor_total': float(cat.valor_total or 0)
                }
                for cat in categorias_stats
            ],
            'productos_mas_vendidos': list(productos_vendidos),
            'rotacion_inventario': rotacion_productos[:10]
        }
        
        return Response(data)

class ValorInventarioViewSet(viewsets.ViewSet):
    """ViewSet para consultas de valor de inventario"""
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    
    def list(self, request):
        """Obtiene el valor del inventario por diferentes criterios"""
        empresa = request.user.empresa
        
        # Valor total por almacén
        valor_por_almacen = []
        for almacen in Almacen.objects.filter(empresa=empresa):
            stocks = Stock.objects.filter(almacen=almacen, empresa=empresa)
            valor_compra = sum(
                stock.cantidad * stock.producto.precio_compra
                for stock in stocks
            )
            valor_venta = sum(
                stock.cantidad * stock.producto.precio_venta
                for stock in stocks
            )
            
            valor_por_almacen.append({
                'almacen': almacen.nombre,
                'valor_total': float(valor_compra),  # Usar precio de compra como valor total
                'valor_compra': float(valor_compra),
                'valor_venta': float(valor_venta),
                'utilidad_potencial': float(valor_venta - valor_compra)
            })
        
        # Valor total por categoría
        valor_por_categoria = []
        for categoria in Categoria.objects.filter(empresa=empresa):
            productos = Producto.objects.filter(categoria=categoria, empresa=empresa, is_active=True)
            valor_compra = sum(
                producto.stock_total * producto.precio_compra
                for producto in productos
            )
            valor_venta = sum(
                producto.stock_total * producto.precio_venta
                for producto in productos
            )
            
            valor_por_categoria.append({
                'categoria': categoria.nombre,
                'valor_total': float(valor_compra),  # Usar precio de compra como valor total
                'valor_compra': float(valor_compra),
                'valor_venta': float(valor_venta),
                'utilidad_potencial': float(valor_venta - valor_compra)
            })
        
        # Resumen general
        total_productos = Producto.objects.filter(empresa=empresa, is_active=True)
        valor_total_compra = sum(
            producto.stock_total * producto.precio_compra
            for producto in total_productos
        )
        valor_total_venta = sum(
            producto.stock_total * producto.precio_venta
            for producto in total_productos
        )
        
        data = {
            'resumen_general': {
                'valor_total': float(valor_total_compra),  # Usar precio de compra como valor total
                'valor_total_compra': float(valor_total_compra),
                'valor_total_venta': float(valor_total_venta),
                'utilidad_potencial': float(valor_total_venta - valor_total_compra)
            },
            'valor_por_almacen': valor_por_almacen,
            'valor_por_categoria': valor_por_categoria
        }
        
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def reporte_detallado(self, request):
        """Obtiene el reporte detallado del valor del inventario por productos"""
        empresa = request.user.empresa
        
        # Obtener productos con stock
        productos = Producto.objects.filter(
            empresa=empresa, 
            is_active=True,
            stock_total__gt=0
        )
        
        items = []
        for producto in productos:
            valor_total = float(producto.stock_total * producto.precio_compra)
            items.append({
                'producto__nombre': producto.nombre,
                'producto__codigo': producto.sku,
                'descripcion': producto.descripcion,
                'categoria': producto.categoria.nombre if producto.categoria else 'Sin categoría',
                'cantidad_total': float(producto.stock_total),
                'producto__unidad_medida': producto.unidad_medida,
                'estado': producto.get_estado_display(),
                'costo_promedio': float(producto.precio_compra),
                'valor_total': valor_total
            })
        
        return Response(items)

# Clases para importar/exportar productos
class ImportarProductosAPIView(views.APIView):
    """APIView para importar productos desde archivo Excel"""
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        """Importa productos desde archivo Excel"""
        archivo = request.FILES.get('file')  # Cambiado de 'archivo' a 'file'
        if not archivo:
            return Response({'error': 'No se proporcionó archivo'}, status=400)
        
        try:
            df = pd.read_excel(archivo)

            df.columns = df.columns.str.strip().str.lower()
            existing_cols = set(df.columns)
            COLUMN_MAP = {}
            if 'nombre' not in existing_cols:
                for alias in ['producto', 'name', 'item']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'nombre'
                        break
            if 'sku' not in existing_cols:
                for alias in ['código', 'codigo', 'cod', 'code']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'sku'
                        break
            if 'precio_compra' not in existing_cols:
                for alias in ['precio compra', 'costo', 'precio_costo', 'cost']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'precio_compra'
                        break
            if 'precio_venta' not in existing_cols:
                for alias in ['precio venta', 'venta', 'price', 'precio']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'precio_venta'
                        break
            if 'categoria' not in existing_cols:
                for alias in ['categoría', 'category']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'categoria'
                        break
            if 'almacen' not in existing_cols:
                for alias in ['almacén', 'warehouse']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'almacen'
                        break
            if 'stock_inicial' not in existing_cols:
                for alias in ['stock', 'stock inicial', 'cantidad']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'stock_inicial'
                        break
            if 'stock_minimo' not in existing_cols:
                for alias in ['stock mínimo', 'stock minimo']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'stock_minimo'
                        break
            if 'stock_maximo' not in existing_cols:
                for alias in ['stock máximo', 'stock maximo']:
                    if alias in existing_cols:
                        COLUMN_MAP[alias] = 'stock_maximo'
                        break

            if COLUMN_MAP:
                df.rename(columns=COLUMN_MAP, inplace=True)

            columnas_requeridas = ['sku', 'nombre', 'precio_compra', 'precio_venta']

            if 'precio_compra' not in df.columns and 'precio_venta' in df.columns:
                df['precio_compra'] = df['precio_venta']
            if 'precio_venta' not in df.columns and 'precio_compra' in df.columns:
                df['precio_venta'] = df['precio_compra']

            missing = [c for c in columnas_requeridas if c not in df.columns]
            if missing:
                return Response({
                    'error': f'Columnas requeridas faltantes: {", ".join(missing)}. '
                             f'Columnas encontradas: {", ".join(df.columns)}'
                }, status=400)
            
            productos_creados = 0
            categorias_creadas = []
            almacenes_creados = []
            errores = []
            
            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        sku_val = row['sku'] if not isinstance(row['sku'], pd.Series) else row['sku'].iloc[0]
                        nombre_val = row['nombre'] if not isinstance(row['nombre'], pd.Series) else row['nombre'].iloc[0]

                        if pd.isna(sku_val) or not str(sku_val).strip():
                            if pd.isna(nombre_val) or not str(nombre_val).strip():
                                continue
                            errores.append(f"Fila {index + 2}: SKU es requerido")
                            continue

                        if pd.isna(nombre_val) or not str(nombre_val).strip():
                            errores.append(f"Fila {index + 2}: Nombre es requerido")
                            continue

                        def parse_price(val):
                            if pd.isna(val):
                                return 0
                            s = str(val).strip()
                            for ch in ['$', 'S/', 'US$', 'USD', 'PEN', 'S/.']:
                                s = s.replace(ch, '')
                            s = s.strip()
                            if ',' in s and '.' in s:
                                s = s.replace(',', '')
                            elif ',' in s:
                                s = s.replace(',', '.')
                            try:
                                return float(s)
                            except (ValueError, TypeError):
                                return 0

                        precio_compra = parse_price(row['precio_compra'])
                        if precio_compra <= 0:
                            errores.append(f"Fila {index + 2}: Precio de compra debe ser mayor a 0 (valor: '{row['precio_compra']}')")
                            continue

                        precio_venta = parse_price(row['precio_venta'])
                        if precio_venta <= 0:
                            errores.append(f"Fila {index + 2}: Precio de venta debe ser mayor a 0 (valor: '{row['precio_venta']}')")
                            continue
                        
                        clean_sku = str(sku_val).strip()
                        clean_nombre = str(nombre_val).strip()

                        if Producto.objects.filter(
                            empresa=request.user.empresa,
                            sku=clean_sku
                        ).exists():
                            errores.append(f"Fila {index + 2}: SKU '{clean_sku}' ya existe")
                            continue
                        
                        def safe_str(val, default=''):
                            if isinstance(val, pd.Series):
                                val = val.iloc[0]
                            if pd.isna(val):
                                return default
                            return str(val).strip()

                        categoria = None
                        cat_val = safe_str(row.get('categoria'))
                        if cat_val:
                            categoria, created = Categoria.objects.get_or_create(
                                empresa=request.user.empresa,
                                nombre=cat_val
                            )
                            if created and categoria.nombre not in categorias_creadas:
                                categorias_creadas.append(categoria.nombre)

                        almacen = None
                        alm_val = safe_str(row.get('almacen'))
                        if alm_val:
                            almacen, created = Almacen.objects.get_or_create(
                                empresa=request.user.empresa,
                                nombre=alm_val,
                                defaults={'direccion': f"Dirección automática para {alm_val}"}
                            )
                            if created and almacen.nombre not in almacenes_creados:
                                almacenes_creados.append(almacen.nombre)

                        desc_val = safe_str(row.get('descripcion'))
                        moneda_val = safe_str(row.get('moneda'), 'PEN')
                        if moneda_val not in ('PEN', 'USD'):
                            moneda_val = 'PEN'

                        stock_min = parse_price(row.get('stock_minimo', 0))
                        stock_max = parse_price(row.get('stock_maximo', 0))

                        producto = Producto.objects.create(
                            empresa=request.user.empresa,
                            sku=clean_sku,
                            nombre=clean_nombre,
                            descripcion=desc_val,
                            categoria=categoria,
                            precio_compra=precio_compra,
                            precio_venta=precio_venta,
                            moneda=moneda_val,
                            stock_minimo=stock_min,
                            stock_maximo=stock_max
                        )
                        
                        # Ahora asignar el almacén manualmente SIN triggear el save()
                        if almacen:
                            producto.almacen = almacen
                            producto.save(update_fields=['almacen'])
                        
                        # Crear stock inicial si se proporciona
                        if 'stock_inicial' in row and pd.notna(row['stock_inicial']) and almacen:
                            try:
                                # Usar get_or_create para evitar duplicados
                                stock_obj, created = Stock.objects.get_or_create(
                                    empresa=request.user.empresa,
                                    producto=producto,
                                    almacen=almacen,
                                    defaults={'cantidad': row['stock_inicial']}
                                )
                                if not created:
                                    # Si ya existe, reemplazamos la cantidad (no sumamos)
                                    stock_obj.cantidad = float(row['stock_inicial'])
                                    stock_obj.save()
                                    
                            except Exception as stock_error:
                                errores.append(f"Fila {index + 2}: Error en stock - {str(stock_error)}")
                        
                        # IMPORTANTE: Siempre actualizar el stock_total del producto al final
                        producto.actualizar_stock_total()
                        
                        productos_creados += 1
                        
                    except Exception as e:
                        errores.append(f"Fila {index + 2}: {str(e)}")
            
            return Response({
                'success': True,
                'productos_creados': productos_creados,
                'categorias_creadas': categorias_creadas,
                'almacenes_creados': almacenes_creados,
                'errores': errores
            })
            
        except Exception as e:
            return Response({'error': f'Error procesando archivo: {str(e)}'}, status=500)

def descargar_template(request):
    """Descarga template de Excel para importar productos"""
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Productos"
    
    # Headers
    headers = [
        'sku',
        'nombre',
        'descripcion',
        'categoria',
        'almacen',
        'precio_compra',
        'precio_venta',
        'moneda',
        'stock_minimo',
        'stock_maximo',
        'stock_inicial'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de ejemplo
    ejemplos = [
        [
            'PROD001',
            'Producto de Ejemplo 1',
            'Descripción del producto 1',
            'Electrónicos',
            'Almacén Lima',
            100.00,
            150.00,
            'PEN',
            10,
            100,
            50
        ],
        [
            'PROD002',
            'Producto de Ejemplo 2',
            'Descripción del producto 2',
            'Ropa',
            'Almacén Arequipa',
            50.00,
            80.00,
            'USD',
            5,
            50,
            25
        ],
        [
            'PROD003',
            'Producto de Ejemplo 3',
            'Descripción del producto 3',
            'Hogar',
            'Almacén Principal',
            200.00,
            300.00,
            'PEN',
            20,
            200,
            100
        ]
    ]
    
    for row_num, ejemplo in enumerate(ejemplos, 2):
        for col, valor in enumerate(ejemplo, 1):
            ws.cell(row=row_num, column=col, value=valor)
    
    # Agregar instrucciones
    instrucciones_row = len(ejemplos) + 3
    ws.cell(row=instrucciones_row, column=1, value="INSTRUCCIONES:")
    ws.cell(row=instrucciones_row + 1, column=1, value="• Las categorías que no existan se crearán automáticamente")
    ws.cell(row=instrucciones_row + 2, column=1, value="• Los almacenes que no existan se crearán automáticamente")
    ws.cell(row=instrucciones_row + 3, column=1, value="• Los campos sku, nombre, precio_compra y precio_venta son obligatorios")
    ws.cell(row=instrucciones_row + 4, column=1, value="• La moneda puede ser 'PEN' o 'USD' (por defecto: PEN)")
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="template_productos.xlsx"'
    
    wb.save(response)
    return response

class KardexViewSet(viewsets.ModelViewSet):
    """ViewSet para gestionar el Kardex (movimientos de inventario)"""
    
    serializer_class = MovimientoInventarioSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['producto', 'almacen', 'tipo_movimiento', 'tipo_documento']
    search_fields = ['producto__nombre', 'producto__sku', 'numero_documento', 'observaciones']
    ordering_fields = ['fecha', 'created_at']
    ordering = ['-fecha', '-id']
    
    def get_queryset(self):
        """Filtrar por empresa del usuario"""
        return MovimientoInventario.objects.filter(
            empresa=self.request.user.empresa
        ).select_related('producto', 'almacen')
    
    def get_serializer_class(self):
        if self.action == 'create':
            return MovimientoInventarioCreateSerializer
        return MovimientoInventarioSerializer
    
    @action(detail=False, methods=['get'])
    def kardex_producto(self, request):
        """Obtiene el kardex filtrado específicamente por producto y almacén"""
        serializer = KardexSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        filters = {
            'empresa': request.user.empresa
        }
        
        # Aplicar filtros obligatorios
        if not serializer.validated_data.get('producto_id'):
            return Response(
                {'error': 'El producto_id es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not serializer.validated_data.get('almacen_id'):
            return Response(
                {'error': 'El almacen_id es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        filters['producto_id'] = serializer.validated_data['producto_id']
        filters['almacen_id'] = serializer.validated_data['almacen_id']
        
        queryset = MovimientoInventario.objects.filter(**filters)
        
        # Aplicar filtros opcionales de fecha
        if serializer.validated_data.get('fecha_inicio'):
            queryset = queryset.filter(fecha__gte=serializer.validated_data['fecha_inicio'])
        
        if serializer.validated_data.get('fecha_fin'):
            queryset = queryset.filter(fecha__lte=serializer.validated_data['fecha_fin'])
        
        # Ordenar por fecha y ID para mantener el orden cronológico
        queryset = queryset.order_by('fecha', 'id').select_related('producto', 'almacen')
        
        # Obtener información del producto y almacén
        try:
            producto = Producto.objects.get(
                id=serializer.validated_data['producto_id'],
                empresa=request.user.empresa
            )
            almacen = Almacen.objects.get(
                id=serializer.validated_data['almacen_id'],
                empresa=request.user.empresa
            )
        except (Producto.DoesNotExist, Almacen.DoesNotExist):
            return Response(
                {'error': 'Producto o almacén no encontrado'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Calcular saldo inicial
        saldo_inicial = {'cantidad': 0, 'costo_unitario': 0, 'costo_total': 0}
        
        if queryset.exists():
            primer_movimiento = queryset.first()
            
            # El saldo inicial se calcula restando el efecto del primer movimiento
            if primer_movimiento.cantidad_entrada > 0:
                saldo_inicial['cantidad'] = max(0, 
                    float(primer_movimiento.cantidad_saldo) - float(primer_movimiento.cantidad_entrada)
                )
                saldo_inicial['costo_total'] = max(0, 
                    float(primer_movimiento.costo_saldo) - float(primer_movimiento.costo_total_entrada)
                )
            elif primer_movimiento.cantidad_salida > 0:
                saldo_inicial['cantidad'] = (
                    float(primer_movimiento.cantidad_saldo) + float(primer_movimiento.cantidad_salida)
                )
                saldo_inicial['costo_total'] = (
                    float(primer_movimiento.costo_saldo) + float(primer_movimiento.costo_total_salida)
                )
            
            # Calcular costo unitario inicial
            if saldo_inicial['cantidad'] > 0:
                saldo_inicial['costo_unitario'] = saldo_inicial['costo_total'] / saldo_inicial['cantidad']
        
        # Serializar movimientos
        movimientos_serializer = MovimientoInventarioSerializer(queryset, many=True)
        
        return Response({
            'producto': {
                'id': producto.id,
                'sku': producto.sku,
                'nombre': producto.nombre
            },
            'almacen': {
                'id': almacen.id,
                'nombre': almacen.nombre
            },
            'saldo_inicial': saldo_inicial,
            'movimientos': movimientos_serializer.data,
            'total_movimientos': queryset.count()
        })
    
    @action(detail=False, methods=['get'])
    def resumen_kardex(self, request):
        """Obtiene el resumen del kardex por producto y almacén"""
        empresa = request.user.empresa
        
        # Obtener productos con movimientos
        productos_con_movimientos = MovimientoInventario.objects.filter(
            empresa=empresa
        ).values('producto', 'almacen').distinct()
        
        resumenes = []
        
        for item in productos_con_movimientos:
            producto_id = item['producto']
            almacen_id = item['almacen']
            
            try:
                producto = Producto.objects.get(id=producto_id)
                almacen = Almacen.objects.get(id=almacen_id)
            except (Producto.DoesNotExist, Almacen.DoesNotExist):
                continue
            
            # Obtener stock actual
            stock_info = MovimientoInventario.obtener_stock_actual(
                empresa, producto, almacen
            )
            
            # Obtener totales de entradas y salidas
            movimientos = MovimientoInventario.objects.filter(
                empresa=empresa,
                producto_id=producto_id,
                almacen_id=almacen_id
            )
            
            total_entradas = movimientos.aggregate(
                total=Sum('cantidad_entrada')
            )['total'] or Decimal('0')
            
            total_salidas = movimientos.aggregate(
                total=Sum('cantidad_salida')
            )['total'] or Decimal('0')
            
            ultimo_movimiento = movimientos.first()
            
            resumenes.append({
                'producto': producto,
                'almacen': almacen,
                'stock_actual': stock_info['cantidad'],
                'costo_total': stock_info['costo_total'],
                'costo_promedio': stock_info['costo_promedio'],
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'ultimo_movimiento': ultimo_movimiento.fecha if ultimo_movimiento else None
            })
        
        serializer = KardexResumenSerializer(resumenes, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exporta el kardex a Excel con la nueva estructura mejorada"""
        serializer = KardexSerializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        
        # Validar filtros obligatorios
        if not serializer.validated_data.get('producto_id'):
            return Response(
                {'error': 'El producto_id es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not serializer.validated_data.get('almacen_id'):
            return Response(
                {'error': 'El almacen_id es requerido'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        filters = {
            'empresa': request.user.empresa,
            'producto_id': serializer.validated_data['producto_id'],
            'almacen_id': serializer.validated_data['almacen_id']
        }
        
        queryset = MovimientoInventario.objects.filter(**filters)
        
        if serializer.validated_data.get('fecha_inicio'):
            queryset = queryset.filter(fecha__gte=serializer.validated_data['fecha_inicio'])
        
        if serializer.validated_data.get('fecha_fin'):
            queryset = queryset.filter(fecha__lte=serializer.validated_data['fecha_fin'])
        
        queryset = queryset.order_by('fecha', 'id').select_related('producto', 'almacen')
        
        # Obtener información del producto y almacén
        try:
            producto = Producto.objects.get(
                id=serializer.validated_data['producto_id'],
                empresa=request.user.empresa
            )
            almacen = Almacen.objects.get(
                id=serializer.validated_data['almacen_id'],
                empresa=request.user.empresa
            )
        except (Producto.DoesNotExist, Almacen.DoesNotExist):
            return Response(
                {'error': 'Producto o almacén no encontrado'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Calcular saldo inicial
        saldo_inicial = {'cantidad': 0, 'costo_unitario': 0, 'costo_total': 0}
        
        if queryset.exists():
            primer_movimiento = queryset.first()
            
            if primer_movimiento.cantidad_entrada > 0:
                saldo_inicial['cantidad'] = max(0, 
                    float(primer_movimiento.cantidad_saldo) - float(primer_movimiento.cantidad_entrada)
                )
                saldo_inicial['costo_total'] = max(0, 
                    float(primer_movimiento.costo_saldo) - float(primer_movimiento.costo_total_entrada)
                )
            elif primer_movimiento.cantidad_salida > 0:
                saldo_inicial['cantidad'] = (
                    float(primer_movimiento.cantidad_saldo) + float(primer_movimiento.cantidad_salida)
                )
                saldo_inicial['costo_total'] = (
                    float(primer_movimiento.costo_saldo) + float(primer_movimiento.costo_total_salida)
                )
            
            if saldo_inicial['cantidad'] > 0:
                saldo_inicial['costo_unitario'] = saldo_inicial['costo_total'] / saldo_inicial['cantidad']
        
        # Crear workbook de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kardex"
        
        # Configurar estilos
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # Fuentes
        font_titulo = Font(name='Arial', size=14, bold=True)
        font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        font_subheader = Font(name='Arial', size=10, bold=True)
        font_datos = Font(name='Arial', size=10)
        font_saldo_inicial = Font(name='Arial', size=10, bold=True)
        
        # Rellenos de color
        fill_header = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        fill_entradas = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
        fill_salidas = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')
        fill_saldos = PatternFill(start_color='E5E0EC', end_color='E5E0EC', fill_type='solid')
        fill_saldo_inicial = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        
        # Bordes
        border_thick = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alineación
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        
        # ENCABEZADO DEL REPORTE
        ws.merge_cells('A1:J1')
        ws['A1'] = 'KARDEX DE INVENTARIO'
        ws['A1'].font = font_titulo
        ws['A1'].alignment = align_center
        
        # Información del producto y almacén
        ws['A3'] = 'Producto:'
        ws['B3'] = f"{producto.sku} - {producto.nombre}"
        ws['A4'] = 'Almacén:'
        ws['B4'] = almacen.nombre
        ws['A5'] = 'Fecha de reporte:'
        ws['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        # Rango de fechas si aplica
        if serializer.validated_data.get('fecha_inicio') and serializer.validated_data.get('fecha_fin'):
            ws['A6'] = 'Período:'
            ws['B6'] = f"{serializer.validated_data['fecha_inicio'].strftime('%d/%m/%Y')} al {serializer.validated_data['fecha_fin'].strftime('%d/%m/%Y')}"
            start_row = 8
        else:
            start_row = 7
        
        # HEADERS PRINCIPALES
        # Fila 1 de headers - FECHA (fusionar verticalmente)
        ws.merge_cells(f'A{start_row}:A{start_row+1}')
        ws[f'A{start_row}'] = 'FECHA'
        ws[f'A{start_row}'].font = font_header
        ws[f'A{start_row}'].fill = fill_header
        ws[f'A{start_row}'].alignment = align_center
        ws[f'A{start_row}'].border = border_thick
        
        # ENTRADAS (fusionar horizontalmente solo en la primera fila)
        ws.merge_cells(f'B{start_row}:D{start_row}')
        ws[f'B{start_row}'] = 'ENTRADAS'
        ws[f'B{start_row}'].font = font_header
        ws[f'B{start_row}'].fill = fill_entradas
        ws[f'B{start_row}'].alignment = align_center
        ws[f'B{start_row}'].border = border_thick
        
        # SALIDAS (fusionar horizontalmente solo en la primera fila)
        ws.merge_cells(f'E{start_row}:G{start_row}')
        ws[f'E{start_row}'] = 'SALIDAS'
        ws[f'E{start_row}'].font = font_header
        ws[f'E{start_row}'].fill = fill_salidas
        ws[f'E{start_row}'].alignment = align_center
        ws[f'E{start_row}'].border = border_thick
        
        # SALDOS (fusionar horizontalmente solo en la primera fila)
        ws.merge_cells(f'H{start_row}:J{start_row}')
        ws[f'H{start_row}'] = 'SALDOS'
        ws[f'H{start_row}'].font = font_header
        ws[f'H{start_row}'].fill = fill_saldos
        ws[f'H{start_row}'].alignment = align_center
        ws[f'H{start_row}'].border = border_thick
        
        # Fila 2 de headers (subheaders) - Solo las columnas no fusionadas
        # La columna A (FECHA) ya está fusionada, así que empezamos desde B
        subheaders_info = [
            # (columna, texto, color_fill)
            (2, 'CANTIDAD', fill_entradas),      # B
            (3, 'COSTO UNIT.', fill_entradas),   # C
            (4, 'COSTO TOTAL', fill_entradas),   # D
            (5, 'CANTIDAD', fill_salidas),       # E
            (6, 'COSTO UNIT.', fill_salidas),    # F
            (7, 'COSTO TOTAL', fill_salidas),    # G
            (8, 'CANTIDAD', fill_saldos),        # H
            (9, 'COSTO UNIT.', fill_saldos),     # I
            (10, 'COSTO TOTAL', fill_saldos),    # J
        ]
        
        for col_num, header_text, color_fill in subheaders_info:
            cell = ws.cell(row=start_row+1, column=col_num, value=header_text)
            cell.font = font_subheader
            cell.alignment = align_center
            cell.border = border_thin
            cell.fill = color_fill
        
        # SALDO INICIAL
        current_row = start_row + 2
        if saldo_inicial['cantidad'] > 0 or saldo_inicial['costo_total'] > 0:
            ws[f'A{current_row}'] = 'SALDO INICIAL'
            ws[f'A{current_row}'].font = font_saldo_inicial
            ws[f'A{current_row}'].fill = fill_saldo_inicial
            ws[f'A{current_row}'].alignment = align_center
            
            # Entradas vacías
            for col in range(2, 5):
                cell = ws.cell(row=current_row, column=col, value='-')
                cell.fill = fill_entradas
                cell.alignment = align_center
                cell.border = border_thin
            
            # Salidas vacías
            for col in range(5, 8):
                cell = ws.cell(row=current_row, column=col, value='-')
                cell.fill = fill_salidas
                cell.alignment = align_center
                cell.border = border_thin
            
            # Saldos iniciales
            ws[f'H{current_row}'] = f"{saldo_inicial['cantidad']:.2f}"
            ws[f'I{current_row}'] = f"$ {saldo_inicial['costo_unitario']:.2f}"
            ws[f'J{current_row}'] = f"$ {saldo_inicial['costo_total']:.2f}"
            
            for col in range(8, 11):
                cell = ws.cell(row=current_row, column=col)
                cell.font = font_saldo_inicial
                cell.fill = fill_saldos
                cell.alignment = align_right
                cell.border = border_thin
            
            current_row += 1
        
        # DATOS DE MOVIMIENTOS
        for movimiento in queryset:
            # Fecha y tipo de movimiento
            fecha_str = movimiento.fecha.strftime('%d/%m/%Y')
            if hasattr(movimiento, 'get_tipo_movimiento_display'):
                fecha_str += f"\n{movimiento.get_tipo_movimiento_display()}"
            if movimiento.numero_documento:
                fecha_str += f"\nDoc: {movimiento.numero_documento}"
            
            ws[f'A{current_row}'] = fecha_str
            ws[f'A{current_row}'].font = font_datos
            ws[f'A{current_row}'].alignment = align_center
            ws[f'A{current_row}'].border = border_thin
            
            # ENTRADAS
            if movimiento.cantidad_entrada > 0:
                ws[f'B{current_row}'] = f"{float(movimiento.cantidad_entrada):.2f}"
                ws[f'C{current_row}'] = f"$ {float(movimiento.costo_unitario):.2f}"
                ws[f'D{current_row}'] = f"$ {float(movimiento.costo_total_entrada):.2f}"
            else:
                ws[f'B{current_row}'] = '-'
                ws[f'C{current_row}'] = '-'
                ws[f'D{current_row}'] = '-'
            
            for col in range(2, 5):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_entradas
                cell.alignment = align_right if cell.value != '-' else align_center
                cell.border = border_thin
                cell.font = font_datos
            
            # SALIDAS
            if movimiento.cantidad_salida > 0:
                ws[f'E{current_row}'] = f"{float(movimiento.cantidad_salida):.2f}"
                ws[f'F{current_row}'] = f"$ {float(movimiento.costo_unitario):.2f}"
                ws[f'G{current_row}'] = f"$ {float(movimiento.costo_total_salida):.2f}"
            else:
                ws[f'E{current_row}'] = '-'
                ws[f'F{current_row}'] = '-'
                ws[f'G{current_row}'] = '-'
            
            for col in range(5, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_salidas
                cell.alignment = align_right if cell.value != '-' else align_center
                cell.border = border_thin
                cell.font = font_datos
            
            # SALDOS
            ws[f'H{current_row}'] = f"{float(movimiento.cantidad_saldo):.2f}"
            ws[f'I{current_row}'] = f"$ {float(movimiento.costo_promedio):.2f}"
            ws[f'J{current_row}'] = f"$ {float(movimiento.costo_saldo):.2f}"
            
            for col in range(8, 11):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill_saldos
                cell.alignment = align_right
                cell.border = border_thin
                cell.font = Font(name='Arial', size=10, bold=True)
            
            current_row += 1
        
        # Ajustar ancho de columnas
        column_widths = [20, 12, 12, 15, 12, 12, 15, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Kardex_{producto.sku}_{almacen.nombre.replace(" ", "_")}_{fecha_actual}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'])
    def ajuste_inventario(self, request):
        """Crea un ajuste de inventario (entrada o salida manual)"""
        data = request.data.copy()
        data['tipo_documento'] = 'ajuste'
        
        # Determinar tipo de movimiento basado en la cantidad
        cantidad_entrada = Decimal(str(data.get('cantidad_entrada', 0)))
        cantidad_salida = Decimal(str(data.get('cantidad_salida', 0)))
        
        if cantidad_entrada > 0:
            data['tipo_movimiento'] = 'ajuste_entrada'
        elif cantidad_salida > 0:
            data['tipo_movimiento'] = 'ajuste_salida'
        
        serializer = MovimientoInventarioCreateSerializer(
            data=data, 
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        movimiento = serializer.save()
        
        response_serializer = MovimientoInventarioSerializer(movimiento)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)


# =====================================================
# VIEWS PARA INVENTARIOS SEPARADOS (MATERIAS PRIMAS Y PRODUCTOS TERMINADOS)
# =====================================================

class InventarioMateriasPrimasViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar el inventario de materias primas e insumos.
    Proporciona CRUD completo y acciones adicionales para operaciones específicas.
    Optimizado con cache y queries eficientes.
    """
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['producto', 'almacen', 'lote']
    search_fields = ['producto__nombre', 'producto__sku', 'ubicacion_almacen', 'lote']
    ordering_fields = ['cantidad_disponible', 'costo_unitario_promedio', 'fecha_vencimiento', 'ultima_actualizacion']
    ordering = ['producto__nombre']
    
    def get_queryset(self):
        return InventarioMateriasPrimas.objects.filter(
            empresa=self.request.user.empresa
        ).select_related(
            'producto', 
            'producto__categoria',
            'almacen'
        ).only(
            'id', 'cantidad_disponible', 'cantidad_reservada', 
            'costo_unitario_promedio', 'ubicacion_almacen',
            'stock_minimo', 'stock_maximo', 'lote', 
            'fecha_vencimiento', 'ultima_actualizacion',
            'producto__id', 'producto__nombre', 'producto__sku',
            'producto__unidad_medida', 'producto__categoria__nombre',
            'almacen__id', 'almacen__nombre'
        )
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return InventarioMateriasPrimasCreateSerializer
        return InventarioMateriasPrimasSerializer
    
    def list(self, request, *args, **kwargs):
        """Override list para caching"""
        empresa_id = request.user.empresa_id
        cache_key = make_cache_key('inventario_mp_list', empresa_id)
        
        # Sin filtros, intentar cache
        if not request.query_params:
            cached_data = cache.get(cache_key)
            if cached_data is not None:
                return Response(cached_data)
        
        response = super().list(request, *args, **kwargs)
        
        if not request.query_params:
            cache.set(cache_key, response.data, get_cache_timeout('inventario_mp'))
        
        return response
    
    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)
        cache_manager.invalidate_inventario_mp(self.request.user.empresa_id)
    
    def perform_update(self, serializer):
        serializer.save()
        cache_manager.invalidate_inventario_mp(self.request.user.empresa_id)
    
    def perform_destroy(self, instance):
        instance.delete()
        cache_manager.invalidate_inventario_mp(self.request.user.empresa_id)
    
    @action(detail=False, methods=['get'])
    def stock_bajo(self, request):
        """Obtiene materias primas con stock bajo"""
        inventarios = self.get_queryset().filter(
            cantidad_disponible__lte=F('stock_minimo'),
            stock_minimo__gt=0
        )
        serializer = InventarioMateriasPrimasSerializer(inventarios, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def por_vencer(self, request):
        """Obtiene materias primas próximas a vencer (30 días)"""
        fecha_limite = timezone.now().date() + timedelta(days=30)
        inventarios = self.get_queryset().filter(
            fecha_vencimiento__lte=fecha_limite,
            fecha_vencimiento__isnull=False,
            cantidad_disponible__gt=0
        ).order_by('fecha_vencimiento')
        serializer = InventarioMateriasPrimasSerializer(inventarios, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def alertas(self, request):
        """Obtiene todas las alertas de materias primas"""
        alertas = InventarioMateriasPrimas.obtener_alertas_stock(request.user.empresa)
        serializer = AlertaStockSerializer(alertas, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def entrada_compra(self, request):
        """Registra entrada de materia prima por compra"""
        serializer = EntradaCompraSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        empresa = request.user.empresa
        
        try:
            with transaction.atomic():
                # Registrar entrada en inventario de materias primas
                inventario = InventarioMateriasPrimas.registrar_entrada_compra(
                    empresa=empresa,
                    producto=data['producto'],
                    almacen=data['almacen'],
                    cantidad=data['cantidad'],
                    costo_unitario=data['costo_unitario'],
                    lote=data.get('lote', ''),
                    fecha_vencimiento=data.get('fecha_vencimiento'),
                    ubicacion=data.get('ubicacion', '')
                )
                
                # Registrar movimiento de inventario
                MovimientoInventario.objects.create(
                    empresa=empresa,
                    producto=data['producto'],
                    almacen=data['almacen'],
                    fecha=timezone.now(),
                    tipo_movimiento='entrada_compra',
                    tipo_documento='compra',
                    tipo_inventario='materia_prima',
                    numero_documento=str(data.get('compra_id', '')),
                    cantidad_entrada=data['cantidad'],
                    costo_unitario=data['costo_unitario'],
                    costo_total_entrada=data['cantidad'] * data['costo_unitario'],
                    cantidad_saldo=inventario.cantidad_disponible + inventario.cantidad_reservada,
                    costo_saldo=inventario.valor_inventario,
                    costo_promedio=inventario.costo_unitario_promedio,
                    inventario_anterior=inventario.cantidad_disponible + inventario.cantidad_reservada - data['cantidad'],
                    inventario_actual=inventario.cantidad_disponible + inventario.cantidad_reservada,
                    compra_id=data.get('compra_id'),
                    lote_referencia=data.get('lote', ''),
                    observaciones=f"Entrada por compra ID: {data.get('compra_id', 'N/A')}",
                    created_by=request.user.username
                )
                
                return Response({
                    'success': True,
                    'message': 'Entrada registrada correctamente',
                    'inventario': InventarioMateriasPrimasSerializer(inventario).data
                }, status=status.HTTP_201_CREATED)
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def reservar(self, request, pk=None):
        """Reserva material para producción"""
        inventario = self.get_object()
        cantidad = Decimal(str(request.data.get('cantidad', 0)))
        orden_produccion_id = request.data.get('orden_produccion_id')
        
        if cantidad <= 0:
            return Response({'error': 'La cantidad debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                inventario_actualizado = InventarioMateriasPrimas.reservar_para_produccion(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    cantidad=cantidad
                )
                
                # Registrar movimiento
                MovimientoInventario.objects.create(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    fecha=timezone.now(),
                    tipo_movimiento='reserva_produccion',
                    tipo_documento='orden_produccion',
                    tipo_inventario='materia_prima',
                    numero_documento=str(orden_produccion_id or ''),
                    cantidad_salida=cantidad,
                    costo_unitario=inventario.costo_unitario_promedio,
                    costo_total_salida=cantidad * inventario.costo_unitario_promedio,
                    cantidad_saldo=inventario_actualizado.cantidad_disponible,
                    observaciones=f"Reserva para orden de producción: {orden_produccion_id}",
                    orden_produccion_id=orden_produccion_id,
                    created_by=request.user.username
                )
                
                return Response({
                    'success': True,
                    'message': f'Reservado {cantidad} unidades correctamente',
                    'inventario': InventarioMateriasPrimasSerializer(inventario_actualizado).data
                })
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def liberar_reserva(self, request, pk=None):
        """Libera reserva de material"""
        inventario = self.get_object()
        cantidad = Decimal(str(request.data.get('cantidad', 0)))
        
        if cantidad <= 0:
            return Response({'error': 'La cantidad debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            inventario_actualizado = InventarioMateriasPrimas.liberar_reserva(
                empresa=request.user.empresa,
                producto=inventario.producto,
                almacen=inventario.almacen,
                cantidad=cantidad
            )
            
            return Response({
                'success': True,
                'message': f'Liberado {cantidad} unidades de reserva',
                'inventario': InventarioMateriasPrimasSerializer(inventario_actualizado).data
            })
            
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def ajustar(self, request, pk=None):
        """Realiza un ajuste de inventario"""
        inventario = self.get_object()
        cantidad_ajuste = Decimal(str(request.data.get('cantidad_ajuste', 0)))
        motivo = request.data.get('motivo', '')
        
        if cantidad_ajuste == 0:
            return Response({'error': 'La cantidad de ajuste no puede ser 0'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                inventario_actualizado = InventarioMateriasPrimas.ajustar_inventario(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    cantidad_ajuste=cantidad_ajuste,
                    motivo=motivo
                )
                
                # Registrar movimiento
                tipo_mov = 'ajuste_entrada' if cantidad_ajuste > 0 else 'ajuste_salida'
                MovimientoInventario.objects.create(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    fecha=timezone.now(),
                    tipo_movimiento=tipo_mov,
                    tipo_documento='ajuste',
                    tipo_inventario='materia_prima',
                    cantidad_entrada=cantidad_ajuste if cantidad_ajuste > 0 else 0,
                    cantidad_salida=abs(cantidad_ajuste) if cantidad_ajuste < 0 else 0,
                    costo_unitario=inventario.costo_unitario_promedio,
                    cantidad_saldo=inventario_actualizado.cantidad_disponible,
                    observaciones=f"Ajuste de inventario: {motivo}",
                    created_by=request.user.username
                )
                
                return Response({
                    'success': True,
                    'message': f'Ajuste de {cantidad_ajuste} unidades realizado',
                    'inventario': InventarioMateriasPrimasSerializer(inventario_actualizado).data
                })
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def valor_total(self, request):
        """Obtiene el valor total del inventario de materias primas"""
        inventarios = self.get_queryset()
        
        valor_total = sum(
            float(inv.cantidad_total * inv.costo_unitario_promedio)
            for inv in inventarios
        )
        
        cantidad_total = sum(float(inv.cantidad_total) for inv in inventarios)
        
        return Response({
            'valor_total': valor_total,
            'cantidad_total': cantidad_total,
            'total_items': inventarios.count()
        })


class InventarioProductosTerminadosViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar el inventario de productos terminados.
    Proporciona CRUD completo y acciones adicionales para operaciones específicas.
    """
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['producto', 'almacen', 'lote_produccion']
    search_fields = ['producto__nombre', 'producto__sku', 'ubicacion_almacen', 'lote_produccion']
    ordering_fields = ['cantidad_disponible', 'costo_produccion_unitario', 'fecha_produccion', 'ultima_actualizacion']
    ordering = ['producto__nombre']
    
    def get_queryset(self):
        return InventarioProductosTerminados.objects.filter(
            empresa=self.request.user.empresa
        ).select_related('producto', 'almacen')
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return InventarioProductosTerminadosCreateSerializer
        return InventarioProductosTerminadosSerializer
    
    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)
    
    @action(detail=False, methods=['get'])
    def stock_bajo(self, request):
        """Obtiene productos terminados con stock bajo"""
        inventarios = self.get_queryset().filter(
            cantidad_disponible__lte=F('stock_minimo'),
            stock_minimo__gt=0
        )
        serializer = InventarioProductosTerminadosSerializer(inventarios, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def alertas(self, request):
        """Obtiene todas las alertas de productos terminados"""
        alertas = InventarioProductosTerminados.obtener_alertas_stock(request.user.empresa)
        serializer = AlertaStockSerializer(alertas, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def entrada_produccion(self, request):
        """Registra entrada de producto terminado desde producción"""
        serializer = EntradaProduccionSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        empresa = request.user.empresa
        
        try:
            with transaction.atomic():
                # Registrar entrada en inventario de productos terminados
                inventario = InventarioProductosTerminados.registrar_entrada_produccion(
                    empresa=empresa,
                    producto=data['producto'],
                    almacen=data['almacen'],
                    cantidad=data['cantidad'],
                    costo_produccion=data['costo_produccion'],
                    lote=data.get('lote', ''),
                    orden_produccion_id=data.get('orden_produccion_id'),
                    fecha_produccion=data.get('fecha_produccion'),
                    fecha_vencimiento=data.get('fecha_vencimiento'),
                    ubicacion=data.get('ubicacion', '')
                )
                
                # Registrar movimiento de inventario
                MovimientoInventario.objects.create(
                    empresa=empresa,
                    producto=data['producto'],
                    almacen=data['almacen'],
                    fecha=timezone.now(),
                    tipo_movimiento='entrada_produccion',
                    tipo_documento='recepcion_produccion',
                    tipo_inventario='producto_terminado',
                    numero_documento=str(data.get('orden_produccion_id', '')),
                    cantidad_entrada=data['cantidad'],
                    costo_unitario=data['costo_produccion'],
                    costo_total_entrada=data['cantidad'] * data['costo_produccion'],
                    cantidad_saldo=inventario.cantidad_disponible + inventario.cantidad_reservada,
                    costo_saldo=inventario.valor_inventario,
                    costo_promedio=inventario.costo_produccion_unitario,
                    inventario_anterior=inventario.cantidad_disponible + inventario.cantidad_reservada - data['cantidad'],
                    inventario_actual=inventario.cantidad_disponible + inventario.cantidad_reservada,
                    orden_produccion_id=data.get('orden_produccion_id'),
                    lote_referencia=data.get('lote', ''),
                    observaciones=f"Entrada por producción - Orden: {data.get('orden_produccion_id', 'N/A')}",
                    created_by=request.user.username
                )
                
                return Response({
                    'success': True,
                    'message': 'Entrada de producción registrada correctamente',
                    'inventario': InventarioProductosTerminadosSerializer(inventario).data
                }, status=status.HTTP_201_CREATED)
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def salida_venta(self, request):
        """Registra salida de producto terminado por venta"""
        serializer = SalidaVentaSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        empresa = request.user.empresa
        
        try:
            with transaction.atomic():
                # Procesar salida por venta
                resultado = InventarioProductosTerminados.procesar_venta(
                    empresa=empresa,
                    producto=data['producto'],
                    almacen=data['almacen'],
                    cantidad=data['cantidad'],
                    desde_reserva=data.get('desde_reserva', False)
                )
                
                inventario = resultado['inventario']
                
                # Registrar movimiento de inventario
                MovimientoInventario.objects.create(
                    empresa=empresa,
                    producto=data['producto'],
                    almacen=data['almacen'],
                    fecha=timezone.now(),
                    tipo_movimiento='salida_venta',
                    tipo_documento='venta',
                    tipo_inventario='producto_terminado',
                    numero_documento=str(data.get('venta_id', '')),
                    cantidad_salida=data['cantidad'],
                    costo_unitario=resultado['costo_unitario'],
                    costo_total_salida=resultado['costo_venta'],
                    cantidad_saldo=inventario.cantidad_disponible + inventario.cantidad_reservada,
                    costo_saldo=inventario.valor_inventario,
                    costo_promedio=inventario.costo_produccion_unitario,
                    inventario_anterior=inventario.cantidad_disponible + inventario.cantidad_reservada + data['cantidad'],
                    inventario_actual=inventario.cantidad_disponible + inventario.cantidad_reservada,
                    venta_id=data.get('venta_id'),
                    observaciones=f"Salida por venta ID: {data.get('venta_id', 'N/A')}",
                    created_by=request.user.username
                )
                
                return Response({
                    'success': True,
                    'message': 'Salida por venta registrada correctamente',
                    'inventario': InventarioProductosTerminadosSerializer(inventario).data,
                    'costo_venta': float(resultado['costo_venta'])
                }, status=status.HTTP_201_CREATED)
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'])
    def reservar(self, request, pk=None):
        """Reserva producto para venta pendiente"""
        inventario = self.get_object()
        cantidad = Decimal(str(request.data.get('cantidad', 0)))
        venta_id = request.data.get('venta_id')
        
        if cantidad <= 0:
            return Response({'error': 'La cantidad debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                inventario_actualizado = InventarioProductosTerminados.reservar_para_venta(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    cantidad=cantidad
                )
                
                # Registrar movimiento
                MovimientoInventario.objects.create(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    fecha=timezone.now(),
                    tipo_movimiento='reserva_venta',
                    tipo_documento='venta',
                    tipo_inventario='producto_terminado',
                    numero_documento=str(venta_id or ''),
                    cantidad_salida=cantidad,
                    costo_unitario=inventario.costo_produccion_unitario,
                    costo_total_salida=cantidad * inventario.costo_produccion_unitario,
                    cantidad_saldo=inventario_actualizado.cantidad_disponible,
                    observaciones=f"Reserva para venta: {venta_id}",
                    venta_id=venta_id,
                    created_by=request.user.username
                )
                
                return Response({
                    'success': True,
                    'message': f'Reservado {cantidad} unidades correctamente',
                    'inventario': InventarioProductosTerminadosSerializer(inventario_actualizado).data
                })
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def liberar_reserva(self, request, pk=None):
        """Libera reserva de producto"""
        inventario = self.get_object()
        cantidad = Decimal(str(request.data.get('cantidad', 0)))
        
        if cantidad <= 0:
            return Response({'error': 'La cantidad debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            inventario_actualizado = InventarioProductosTerminados.liberar_reserva(
                empresa=request.user.empresa,
                producto=inventario.producto,
                almacen=inventario.almacen,
                cantidad=cantidad
            )
            
            return Response({
                'success': True,
                'message': f'Liberado {cantidad} unidades de reserva',
                'inventario': InventarioProductosTerminadosSerializer(inventario_actualizado).data
            })
            
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def ajustar(self, request, pk=None):
        """Realiza un ajuste de inventario"""
        inventario = self.get_object()
        cantidad_ajuste = Decimal(str(request.data.get('cantidad_ajuste', 0)))
        motivo = request.data.get('motivo', '')
        
        if cantidad_ajuste == 0:
            return Response({'error': 'La cantidad de ajuste no puede ser 0'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            with transaction.atomic():
                inventario_actualizado = InventarioProductosTerminados.ajustar_inventario(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    cantidad_ajuste=cantidad_ajuste,
                    motivo=motivo
                )
                
                # Registrar movimiento
                tipo_mov = 'ajuste_entrada' if cantidad_ajuste > 0 else 'ajuste_salida'
                MovimientoInventario.objects.create(
                    empresa=request.user.empresa,
                    producto=inventario.producto,
                    almacen=inventario.almacen,
                    fecha=timezone.now(),
                    tipo_movimiento=tipo_mov,
                    tipo_documento='ajuste',
                    tipo_inventario='producto_terminado',
                    cantidad_entrada=cantidad_ajuste if cantidad_ajuste > 0 else 0,
                    cantidad_salida=abs(cantidad_ajuste) if cantidad_ajuste < 0 else 0,
                    costo_unitario=inventario.costo_produccion_unitario,
                    cantidad_saldo=inventario_actualizado.cantidad_disponible,
                    observaciones=f"Ajuste de inventario: {motivo}",
                    created_by=request.user.username
                )
                
                return Response({
                    'success': True,
                    'message': f'Ajuste de {cantidad_ajuste} unidades realizado',
                    'inventario': InventarioProductosTerminadosSerializer(inventario_actualizado).data
                })
                
        except ValidationError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def valor_total(self, request):
        """Obtiene el valor total del inventario de productos terminados"""
        inventarios = self.get_queryset()
        
        valor_costo = sum(
            float(inv.cantidad_total * inv.costo_produccion_unitario)
            for inv in inventarios
        )
        
        valor_venta = sum(
            float(inv.cantidad_disponible * inv.precio_venta_sugerido)
            for inv in inventarios
        )
        
        cantidad_total = sum(float(inv.cantidad_total) for inv in inventarios)
        
        return Response({
            'valor_costo': valor_costo,
            'valor_venta_potencial': valor_venta,
            'margen_potencial': valor_venta - valor_costo,
            'cantidad_total': cantidad_total,
            'total_items': inventarios.count()
        })


class ResumenInventariosSeparadosView(views.APIView):
    """
    Vista para obtener resumen de ambos inventarios (materias primas y productos terminados).
    """
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    
    def get(self, request):
        empresa = request.user.empresa
        
        # Resumen de materias primas
        materias_primas = InventarioMateriasPrimas.objects.filter(empresa=empresa)
        total_mp = materias_primas.count()
        valor_mp = sum(
            float(inv.cantidad_total * inv.costo_unitario_promedio)
            for inv in materias_primas
        )
        mp_stock_bajo = materias_primas.filter(
            cantidad_disponible__lte=F('stock_minimo'),
            stock_minimo__gt=0
        ).count()
        
        # Materias primas por vencer
        fecha_limite = timezone.now().date() + timedelta(days=30)
        mp_por_vencer = materias_primas.filter(
            fecha_vencimiento__lte=fecha_limite,
            fecha_vencimiento__isnull=False,
            cantidad_disponible__gt=0
        ).count()
        
        # Resumen de productos terminados
        productos_terminados = InventarioProductosTerminados.objects.filter(empresa=empresa)
        total_pt = productos_terminados.count()
        valor_pt = sum(
            float(inv.cantidad_total * inv.costo_produccion_unitario)
            for inv in productos_terminados
        )
        pt_stock_bajo = productos_terminados.filter(
            cantidad_disponible__lte=F('stock_minimo'),
            stock_minimo__gt=0
        ).count()
        valor_venta = sum(
            float(inv.cantidad_disponible * inv.precio_venta_sugerido)
            for inv in productos_terminados
        )
        
        # Obtener todas las alertas
        alertas_mp = InventarioMateriasPrimas.obtener_alertas_stock(empresa)
        alertas_pt = InventarioProductosTerminados.obtener_alertas_stock(empresa)
        todas_alertas = alertas_mp + alertas_pt
        
        data = {
            'total_materias_primas': total_mp,
            'valor_total_materias_primas': valor_mp,
            'materias_primas_stock_bajo': mp_stock_bajo,
            'materias_primas_por_vencer': mp_por_vencer,
            'total_productos_terminados': total_pt,
            'valor_total_productos_terminados': valor_pt,
            'productos_terminados_stock_bajo': pt_stock_bajo,
            'valor_venta_potencial': valor_venta,
            'alertas': todas_alertas
        }
        
        serializer = ResumenInventarioSeparadoSerializer(data)
        return Response(serializer.data)


class ValidarStockProduccionView(views.APIView):
    """
    Vista para validar si hay stock suficiente de materias primas para una producción.
    """
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    
    def post(self, request):
        """
        Valida stock para una lista de materiales.
        Espera: { "materiales": [{ "producto_id": 1, "almacen_id": 1, "cantidad": 10 }, ...] }
        """
        materiales = request.data.get('materiales', [])
        
        if not materiales:
            return Response({'error': 'Debe proporcionar lista de materiales'}, status=status.HTTP_400_BAD_REQUEST)
        
        empresa = request.user.empresa
        resultados = []
        puede_producir = True
        
        for material in materiales:
            producto_id = material.get('producto_id')
            almacen_id = material.get('almacen_id')
            cantidad_requerida = Decimal(str(material.get('cantidad', 0)))
            
            try:
                producto = Producto.objects.get(id=producto_id, empresa=empresa)
                almacen = Almacen.objects.get(id=almacen_id, empresa=empresa)
                
                try:
                    inventario = InventarioMateriasPrimas.objects.get(
                        empresa=empresa,
                        producto=producto,
                        almacen=almacen
                    )
                    disponible = inventario.cantidad_disponible
                    suficiente = disponible >= cantidad_requerida
                except InventarioMateriasPrimas.DoesNotExist:
                    disponible = Decimal('0')
                    suficiente = False
                
                if not suficiente:
                    puede_producir = False
                
                resultados.append({
                    'producto_id': producto_id,
                    'producto_nombre': producto.nombre,
                    'almacen_id': almacen_id,
                    'almacen_nombre': almacen.nombre,
                    'cantidad_requerida': float(cantidad_requerida),
                    'cantidad_disponible': float(disponible),
                    'suficiente': suficiente,
                    'faltante': float(max(0, cantidad_requerida - disponible))
                })
                
            except Producto.DoesNotExist:
                resultados.append({
                    'producto_id': producto_id,
                    'error': 'Producto no encontrado',
                    'suficiente': False
                })
                puede_producir = False
            except Almacen.DoesNotExist:
                resultados.append({
                    'almacen_id': almacen_id,
                    'error': 'Almacén no encontrado',
                    'suficiente': False
                })
                puede_producir = False
        
        return Response({
            'puede_producir': puede_producir,
            'validacion_materiales': resultados
        })


# ========================================
# CARGA MASIVA DE INVENTARIO
# ========================================

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
import tempfile
import os


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def descargar_plantilla_inventario(request):
    """
    Descarga la plantilla Excel mejorada para carga masiva de inventario.
    
    La plantilla incluye:
    - Instrucciones de uso
    - Template con headers correctos
    - Ejemplos de materias primas y productos terminados
    - Referencia de categorías válidas
    """
    from .utils.carga_masiva import generar_plantilla_excel
    
    try:
        output = generar_plantilla_excel()
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_carga_inventario.xlsx'
        return response
    
    except Exception as e:
        return Response(
            {'error': f'Error al generar plantilla: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def cargar_inventario_masivo(request):
    """
    Procesa un archivo Excel para carga masiva de inventario.
    
    El sistema detecta automáticamente el tipo de producto basándose en la categoría:
    - Categorías MP/Materia Prima/Insumo → InventarioMateriasPrimas
    - Otras categorías → InventarioProductosTerminados
    
    Request:
        archivo: Archivo Excel (.xlsx)
        
    Response:
        {
            "success": true/false,
            "resumen": {
                "filas_procesadas": 10,
                "productos_validos": 8,
                "materias_primas": 3,
                "productos_terminados": 5,
                "errores": 2,
                "warnings": 1
            },
            "errores": [...],
            "warnings": [...],
            "productos_creados": [...]
        }
    """
    from .utils.carga_masiva import procesar_carga_masiva_excel
    
    if 'archivo' not in request.FILES:
        return Response(
            {'error': 'No se envió ningún archivo. Use el campo "archivo".'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    archivo = request.FILES['archivo']
    
    # Validar extensión
    if not archivo.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'El archivo debe ser un Excel (.xlsx o .xls)'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Obtener empresa del usuario
    empresa = request.user.empresa
    if not empresa:
        return Response(
            {'error': 'Usuario sin empresa asignada'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Guardar archivo temporalmente
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in archivo.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        
        # Procesar archivo
        resultado = procesar_carga_masiva_excel(tmp_path, empresa)
        
        # Eliminar archivo temporal
        os.unlink(tmp_path)
        
        # Determinar código de estado
        if resultado['success']:
            return Response(resultado, status=status.HTTP_201_CREATED)
        elif resultado['resumen']['errores'] > 0:
            return Response(resultado, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(resultado, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    except Exception as e:
        # Limpiar archivo temporal si existe
        if 'tmp_path' in locals():
            try:
                os.unlink(tmp_path)
            except:
                pass
        
        return Response(
            {'error': f'Error al procesar archivo: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def validar_archivo_inventario(request):
    """
    Valida un archivo Excel sin crear productos.
    Útil para preview antes de la carga definitiva.
    
    Request:
        archivo: Archivo Excel (.xlsx)
        
    Response:
        Mismo formato que cargar_inventario_masivo pero sin crear productos
    """
    import pandas as pd
    from .utils.carga_masiva import detectar_tipo_producto, validar_fila
    
    if 'archivo' not in request.FILES:
        return Response(
            {'error': 'No se envió ningún archivo'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    archivo = request.FILES['archivo']
    
    try:
        df = pd.read_excel(archivo, sheet_name='Template Productos')
        df.columns = df.columns.str.lower().str.strip().str.replace(' ', '_')
        
        resultado = {
            'success': True,
            'resumen': {
                'filas_encontradas': len(df),
                'materias_primas': 0,
                'productos_terminados': 0,
                'errores': 0,
                'warnings': 0
            },
            'errores': [],
            'warnings': [],
            'preview': []
        }
        
        for idx, row in df.iterrows():
            numero_fila = idx + 2
            fila = row.to_dict()
            
            if pd.isna(fila.get('sku')) and pd.isna(fila.get('nombre')):
                continue
            
            tipo = detectar_tipo_producto(fila.get('categoria'))
            es_valido, errores, warnings = validar_fila(fila, numero_fila, tipo)
            
            resultado['errores'].extend(errores)
            resultado['warnings'].extend(warnings)
            
            if tipo == 'RAW':
                resultado['resumen']['materias_primas'] += 1
            else:
                resultado['resumen']['productos_terminados'] += 1
            
            resultado['preview'].append({
                'fila': numero_fila,
                'sku': fila.get('sku'),
                'nombre': fila.get('nombre'),
                'tipo_detectado': 'Materia Prima' if tipo == 'RAW' else 'Producto Terminado',
                'valido': es_valido
            })
        
        resultado['resumen']['errores'] = len(resultado['errores'])
        resultado['resumen']['warnings'] = len(resultado['warnings'])
        resultado['success'] = resultado['resumen']['errores'] == 0
        
        return Response(resultado)
        
    except Exception as e:
        return Response(
            {'error': f'Error al validar archivo: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )