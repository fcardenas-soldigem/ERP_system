from rest_framework import generics
from .models import Venta, Cliente, Factura, OrdenVenta, DetalleVenta, PagoVenta
from .serializers import (
    VentaSerializer, ClienteSerializer, FacturaSerializer,
    OrdenVentaSerializer, DetalleVentaSerializer, PagoVentaSerializer
)
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, F, Q
from django.db.models.functions import Coalesce as CoalesceFunc
from django.utils import timezone
from django.http import HttpResponse
from django.shortcuts import render
from rest_framework import viewsets, permissions
from rest_framework.permissions import IsAuthenticated
from apps.core.permissions import HasEmpresaPermission
from django.db import transaction
from apps.inventario.models import Stock
from apps.inventario.models.producto import Producto
from rest_framework.decorators import action
from rest_framework import serializers
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.utils import get_column_letter
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
import os
from decimal import Decimal, InvalidOperation
from rest_framework.pagination import PageNumberPagination
from django.db import models
from openpyxl import Workbook
from django.shortcuts import get_object_or_404

class VentaPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class VentaListCreateView(generics.ListCreateAPIView):
    queryset = Venta.objects.all()
    serializer_class = VentaSerializer
    pagination_class = VentaPagination

class VentaRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Venta.objects.all()
    serializer_class = VentaSerializer

class ResumenVentasMensual(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Lógica para el resumen mensual
        return Response({
            'total_ventas': 0,
            'total_clientes': 0,
            'ventas_mes': []
        })

def index_ventas(request):
    return HttpResponse("Bienvenido a las Ventas - Ruta principal (/api/ventas/).")

def resumen_mensual(request):
    return HttpResponse("Resumen Mensual de Ventas")

class ClienteViewSet(viewsets.ModelViewSet):
    serializer_class = ClienteSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        if not self.request.user.empresa:
            return Cliente.objects.none()
        queryset = Cliente.objects.filter(empresa=self.request.user.empresa)
        
        # Búsqueda por nombre o documento
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                models.Q(nombre__icontains=search) |
                models.Q(documento__icontains=search) |
                models.Q(email__icontains=search)
            )
        
        return queryset.order_by('-fecha_registro')

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

    def perform_update(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['post'])
    def importar_clientes(self, request):
        """
        Importar clientes masivamente desde archivo CSV o Excel
        """
        if 'archivo' not in request.FILES:
            return Response({
                'success': False,
                'error': 'Debe proporcionar un archivo'
            }, status=status.HTTP_400_BAD_REQUEST)

        archivo = request.FILES['archivo']
        
        # Validar extensión del archivo
        extension = archivo.name.split('.')[-1].lower()
        if extension not in ['csv', 'xlsx', 'xls']:
            return Response({
                'success': False,
                'error': 'Formato de archivo no soportado. Use CSV o Excel (.xlsx, .xls)'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            import pandas as pd
            from django.db import transaction
            
            # Leer archivo según extensión
            if extension == 'csv':
                df = pd.read_csv(archivo)
            else:
                df = pd.read_excel(archivo)

            # Validar columnas requeridas
            columnas_requeridas = ['nombre', 'documento', 'tipo_documento']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                return Response({
                    'success': False,
                    'error': f'Columnas faltantes: {", ".join(columnas_faltantes)}',
                    'columnas_requeridas': columnas_requeridas,
                    'columnas_opcionales': ['direccion', 'telefono', 'email'],
                    'columnas_encontradas': list(df.columns)
                }, status=status.HTTP_400_BAD_REQUEST)

            # Procesar importación
            exitosos = []
            errores = []
            duplicados = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Validar datos requeridos
                        nombre = str(row['nombre']).strip()
                        documento = str(row['documento']).strip()
                        tipo_documento = str(row['tipo_documento']).strip().lower()

                        if not nombre or nombre == 'nan':
                            errores.append(f'Fila {index + 2}: Nombre es requerido')
                            continue

                        if not documento or documento == 'nan':
                            errores.append(f'Fila {index + 2}: Documento es requerido')
                            continue

                        if tipo_documento not in ['dni', 'ruc']:
                            errores.append(f'Fila {index + 2}: tipo_documento debe ser "dni" o "ruc"')
                            continue

                        # Validar formato de documento
                        if not documento.isdigit():
                            errores.append(f'Fila {index + 2}: Documento debe contener solo números')
                            continue

                        if tipo_documento == 'dni' and len(documento) != 8:
                            errores.append(f'Fila {index + 2}: DNI debe tener 8 dígitos')
                            continue

                        if tipo_documento == 'ruc' and len(documento) != 11:
                            errores.append(f'Fila {index + 2}: RUC debe tener 11 dígitos')
                            continue

                        # Verificar si ya existe
                        if Cliente.objects.filter(
                            empresa=request.user.empresa,
                            documento=documento
                        ).exists():
                            duplicados.append(f'Fila {index + 2}: Cliente con documento {documento} ya existe')
                            continue

                        # Preparar datos opcionales
                        direccion = str(row.get('direccion', '')).strip() if 'direccion' in row and pd.notna(row.get('direccion')) else ''
                        telefono = str(row.get('telefono', '')).strip() if 'telefono' in row and pd.notna(row.get('telefono')) else ''
                        email = str(row.get('email', '')).strip() if 'email' in row and pd.notna(row.get('email')) else ''

                        # Validar teléfono si se proporciona
                        if telefono and telefono != 'nan':
                            if not telefono.isdigit() or len(telefono) != 9:
                                errores.append(f'Fila {index + 2}: Teléfono debe tener 9 dígitos numéricos')
                                continue

                        # Validar email si se proporciona
                        if email and email != 'nan':
                            import re
                            if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
                                errores.append(f'Fila {index + 2}: Email inválido')
                                continue

                        # Crear cliente
                        cliente = Cliente.objects.create(
                            empresa=request.user.empresa,
                            nombre=nombre,
                            documento=documento,
                            tipo_documento=tipo_documento,
                            direccion=direccion or '',
                            telefono=telefono or '',
                            email=email or '',
                            activo=True
                        )

                        exitosos.append({
                            'fila': index + 2,
                            'nombre': cliente.nombre,
                            'documento': cliente.documento,
                            'tipo_documento': cliente.tipo_documento
                        })

                    except Exception as e:
                        errores.append(f'Fila {index + 2}: Error inesperado - {str(e)}')

            return Response({
                'success': True,
                'mensaje': f'Importación completada: {len(exitosos)} exitosos, {len(errores)} errores, {len(duplicados)} duplicados',
                'estadisticas': {
                    'exitosos': len(exitosos),
                    'errores': len(errores),
                    'duplicados': len(duplicados),
                    'total_filas': len(df)
                },
                'detalles_exitosos': exitosos,
                'detalles_errores': errores,
                'detalles_duplicados': duplicados
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al procesar archivo: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def descargar_plantilla_excel(self, request):
        """
        Descargar plantilla Excel para importación masiva de clientes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from django.http import HttpResponse
            import io

            # Crear libro de trabajo
            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')

            # Encabezados
            headers = [
                ('A1', 'nombre', 'Nombre/Razón Social'),
                ('B1', 'documento', 'Documento'),
                ('C1', 'tipo_documento', 'Tipo Documento'),
                ('D1', 'direccion', 'Dirección'),
                ('E1', 'telefono', 'Teléfono'),
                ('F1', 'email', 'Email')
            ]

            # Aplicar encabezados con estilo
            for cell_coord, field, display_name in headers:
                cell = ws[cell_coord]
                cell.value = display_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Datos de ejemplo
            ejemplos = [
                ['Juan Pérez García', '12345678', 'dni', 'Av. Principal 123 - Lima', '987654321', 'juan.perez@email.com'],
                ['María González López', '87654321', 'dni', 'Jr. Los Rosales 456 - San Isidro', '965874123', 'maria.gonzalez@email.com'],
                ['Carlos Mendoza Silva', '45678912', 'dni', 'Calle Las Flores 789 - Miraflores', '932147856', 'carlos.mendoza@email.com'],
                ['Empresa Constructora SAC', '20123456789', 'ruc', 'Av. Javier Prado 1234 - San Borja', '987123456', 'contacto@constructora.com'],
                ['Comercializadora Norte EIRL', '20987654321', 'ruc', 'Jr. Comercio 567 - Cercado de Lima', '987789123', 'ventas@comercializadora.pe'],
                ['Ana Rodríguez Torres', '78912345', 'dni', 'Av. Universitaria 890 - Los Olivos', '945678321', 'ana.rodriguez@email.com'],
                ['Tech Solutions S.A.C.', '20456789123', 'ruc', 'Calle Innovación 321 - Surco', '987456789', 'info@techsolutions.pe'],
                ['Pedro Castillo Vargas', '32165498', 'dni', 'Jr. Independencia 654 - Breña', '912345678', 'pedro.castillo@email.com'],
                ['Distribuidora Central LTDA', '20654321987', 'ruc', 'Av. Central 987 - La Victoria', '987321654', 'admin@distribuidora.com'],
                ['Lucía Fernández Castro', '65432187', 'dni', 'Calle Los Cedros 147 - San Miguel', '923456789', 'lucia.fernandez@email.com']
            ]

            # Agregar datos de ejemplo
            for i, ejemplo in enumerate(ejemplos, start=2):
                for j, valor in enumerate(ejemplo, start=1):
                    cell = ws.cell(row=i, column=j)
                    cell.value = valor
                    cell.border = border

            # Ajustar ancho de columnas
            column_widths = [30, 15, 18, 40, 15, 35]  # Anchos en caracteres
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

            # Crear hoja de instrucciones
            ws_instrucciones = wb.create_sheet("Instrucciones")
            
            # Instrucciones detalladas
            instrucciones = [
                ["📋 INSTRUCCIONES PARA IMPORTAR CLIENTES", ""],
                ["", ""],
                ["1. COLUMNAS REQUERIDAS:", ""],
                ["   • nombre: Nombre completo o razón social", ""],
                ["   • documento: DNI (8 dígitos) o RUC (11 dígitos)", ""],
                ["   • tipo_documento: Escriba 'dni' o 'ruc'", ""],
                ["", ""],
                ["2. COLUMNAS OPCIONALES:", ""],
                ["   • direccion: Dirección completa", ""],
                ["   • telefono: Número de 9 dígitos (solo números)", ""],
                ["   • email: Correo electrónico válido", ""],
                ["", ""],
                ["3. VALIDACIONES IMPORTANTES:", ""],
                ["   • DNI debe tener exactamente 8 dígitos", ""],
                ["   • RUC debe tener exactamente 11 dígitos", ""],
                ["   • Teléfono debe tener exactamente 9 dígitos", ""],
                ["   • No puede haber documentos duplicados", ""],
                ["", ""],
                ["4. FORMATOS SOPORTADOS:", ""],
                ["   • Excel (.xlsx, .xls)", ""],
                ["   • CSV", ""],
                ["", ""],
                ["5. CÓMO USAR:", ""],
                ["   • Use la hoja 'Clientes' como plantilla", ""],
                ["   • Modifique o agregue sus datos", ""],
                ["   • Elimine las filas de ejemplo si no las necesita", ""],
                ["   • Guarde el archivo y súbalo al sistema", ""],
                ["", ""],
                ["💡 CONSEJOS:", ""],
                ["   • Mantenga los encabezados en la primera fila", ""],
                ["   • No deje celdas vacías en campos requeridos", ""],
                ["   • Revise los datos antes de importar", ""],
                ["   • Los duplicados serán omitidos automáticamente", ""]
            ]

            # Agregar instrucciones a la hoja
            for i, (instruccion, _) in enumerate(instrucciones, start=1):
                cell = ws_instrucciones.cell(row=i, column=1)
                cell.value = instruccion
                if "INSTRUCCIONES" in instruccion:
                    cell.font = Font(bold=True, size=14, color="4472C4")
                elif instruccion.endswith(":"):
                    cell.font = Font(bold=True, color="2F5597")
                elif instruccion.startswith("   •"):
                    cell.font = Font(color="666666")

            # Ajustar ancho de columna en instrucciones
            ws_instrucciones.column_dimensions['A'].width = 60

            # Crear respuesta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="plantilla_clientes.xlsx"'

            # Guardar el archivo en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response.write(output.getvalue())

            return response

        except Exception as e:
            return Response({
                'success': False,
                'error': f'Error al generar plantilla Excel: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class FacturaViewSet(viewsets.ModelViewSet):
    queryset = Factura.objects.all()
    serializer_class = FacturaSerializer

class OrdenVentaViewSet(viewsets.ModelViewSet):
    serializer_class = OrdenVentaSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        if not self.request.user.empresa:
            return OrdenVenta.objects.none()
        return OrdenVenta.objects.filter(empresa=self.request.user.empresa)

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

class VentaViewSet(viewsets.ModelViewSet):
    queryset = Venta.objects.all()
    serializer_class = VentaSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    pagination_class = VentaPagination
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        queryset = self.queryset.filter(empresa=self.request.user.empresa)
        
        # Filtros
        estado = self.request.query_params.get('estado', None)
        tipo_venta = self.request.query_params.get('tipo_venta', None)
        metodo_pago = self.request.query_params.get('metodo_pago', None)
        cliente = self.request.query_params.get('cliente', None)
        fecha_inicio = self.request.query_params.get('fecha_inicio', None)
        fecha_fin = self.request.query_params.get('fecha_fin', None)

        if estado:
            queryset = queryset.filter(estado=estado)
        if tipo_venta:
            queryset = queryset.filter(tipo_venta=tipo_venta)
        if metodo_pago:
            queryset = queryset.filter(metodo_pago=metodo_pago)
        if cliente:
            queryset = queryset.filter(cliente_id=cliente)
        if fecha_inicio:
            queryset = queryset.filter(fecha_emision__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha_emision__lte=fecha_fin)

        return queryset.order_by('-fecha_emision')

    def perform_create(self, serializer):
        print(f"DEBUG VIEWSET: Iniciando perform_create")
        print(f"DEBUG VIEWSET: Usuario: {self.request.user}")
        print(f"DEBUG VIEWSET: Empresa del usuario: {self.request.user.empresa}")
        
        try:
            # Simplificamos - el serializer ya maneja la asignación de empresa
            venta = serializer.save()
            print(f"DEBUG VIEWSET: Venta creada: {venta.numero}")
            
            # Si es una venta a crédito, establecer estado inicial como 'pendiente'
            if venta.tipo_venta in ['credito_30', 'credito_60']:
                print(f"DEBUG VIEWSET: Procesando venta a crédito")
                venta.estado = 'pendiente'
                # Calcular fecha de vencimiento basada en el tipo de crédito
                dias_credito = 30 if venta.tipo_venta == 'credito_30' else 60
                venta.fecha_vencimiento = venta.fecha_emision + timezone.timedelta(days=dias_credito)
                venta.save()

                # Crear registro de pago inicial en estado pendiente
                PagoVenta.objects.create(
                    venta=venta,
                    fecha=venta.fecha_emision,
                    monto=venta.total,
                    metodo_pago='pendiente',
                    notas=f'Pago pendiente para venta a crédito de {dias_credito} días'
                )
                print(f"DEBUG VIEWSET: Registro de pago pendiente creado")
            
            # Si la venta está pagada, actualizar el stock inmediatamente
            if venta.estado == 'pagado':
                print(f"=== LLAMANDO ACTUALIZAR_STOCK DESDE PERFORM_CREATE ===")
                print(f"Venta: {venta.numero}, Estado: {venta.estado}")
                venta.actualizar_stock()
                print(f'Stock actualizado para venta {venta.numero}')
                print(f"=== FIN ACTUALIZAR_STOCK DESDE PERFORM_CREATE ===")
                
            print(f"DEBUG VIEWSET: perform_create completado exitosamente")
        except Exception as e:
            print(f"DEBUG VIEWSET ERROR: {str(e)}")
            import traceback
            print(f"DEBUG VIEWSET TRACEBACK: {traceback.format_exc()}")
            raise

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cambiar_estado(self, request, pk=None):
        venta = self.get_object()
        nuevo_estado = request.data.get('estado')
        
        if nuevo_estado not in ['borrador', 'pendiente', 'pagado', 'anulado']:
            return Response(
                {"detail": "Estado no válido"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                if nuevo_estado == 'anulado':
                    venta.anular_venta()
                else:
                    venta.estado = nuevo_estado
                    if nuevo_estado == 'pagado':
                        venta.actualizar_stock()
                    venta.save()
            
            serializer = self.get_serializer(venta)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'])
    def saldo_pendiente(self, request, pk=None):
        venta = self.get_object()
        saldo = venta.get_saldo_pendiente()
        return Response({'saldo_pendiente': saldo})

    @action(detail=False, methods=['get'])
    def ventas_pendientes(self, request):
        """
        Obtener todas las ventas a crédito que están pendientes de pago
        """
        from apps.core.services import DocumentoService
        
        ventas = self.get_queryset().filter(
            tipo_venta__in=['credito_30', 'credito_60'],
            estado='pendiente'
        ).annotate(
            total_pagado=CoalesceFunc(
                Sum('pagos__monto', filter=~Q(pagos__metodo_pago='pendiente')),
                0,
                output_field=models.DecimalField()
            ),
            saldo_pendiente=F('total') - F('total_pagado')
        ).filter(saldo_pendiente__gt=0)

        # Obtener tipo de cambio del día
        try:
            documento_service = DocumentoService()
            tipo_cambio_data = documento_service.obtener_tipo_cambio()
            tipo_cambio_venta = tipo_cambio_data.get('venta', 3.8)  # fallback
        except:
            tipo_cambio_venta = 3.8  # fallback
        
        # Serializar ventas
        serializer = self.get_serializer(ventas, many=True)
        ventas_data = serializer.data
        
        # Agregar conversiones y total general
        total_general_pen = 0
        
        for venta_data in ventas_data:
            moneda = venta_data.get('moneda', 'PEN')
            saldo_pendiente = float(venta_data.get('saldo_pendiente', 0))
            total = float(venta_data.get('total', 0))
            
            # Convertir a PEN para el total general
            if moneda == 'USD':
                saldo_pendiente_pen = saldo_pendiente * tipo_cambio_venta
                total_pen = total * tipo_cambio_venta
            else:
                saldo_pendiente_pen = saldo_pendiente
                total_pen = total
            
            total_general_pen += saldo_pendiente_pen
            
            # Agregar campos de conversión
            venta_data['saldo_pendiente_pen'] = round(saldo_pendiente_pen, 2)
            venta_data['total_pen'] = round(total_pen, 2)
            venta_data['tipo_cambio'] = tipo_cambio_venta
        
        return Response({
            'ventas': ventas_data,
            'total_general_pen': round(total_general_pen, 2),
            'tipo_cambio': tipo_cambio_venta
        })

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request, *args, **kwargs):
        try:
            ventas = self.get_queryset()
            if not ventas.exists():
                return Response(
                    {'error': 'No hay ventas para exportar'},
                    status=status.HTTP_404_NOT_FOUND
                )

            datos = []
            for venta in ventas:
                for detalle in venta.detalles.all():
                    datos.append({
                        'Número': venta.numero,
                        'Fecha': venta.fecha_emision.strftime('%Y-%m-%d') if venta.fecha_emision else '',
                        'Cliente': venta.cliente.nombre if venta.cliente else '',
                        'Producto': detalle.producto.nombre if detalle.producto else '',
                        'Cantidad': float(detalle.cantidad),
                        'Precio Unitario': float(detalle.precio_unitario),
                        'IGV Incluido': 'Sí' if venta.igv_incluido else 'No',
                        'Subtotal': float(venta.subtotal),
                        'IGV': float(venta.igv),
                        'Total': float(venta.total),
                        'Estado': venta.estado.upper() if venta.estado else '',
                        'Método de Pago': venta.metodo_pago.upper() if venta.metodo_pago else ''
                    })

            df = pd.DataFrame(datos)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Ventas', index=False)
                worksheet = writer.sheets['Ventas']
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(str(col))
                    ) + 2
                    worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)
            output.seek(0)
            filename = f'Ventas_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        except Exception as e:
            print(f"Error en exportar_excel: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['POST'], parser_classes=[MultiPartParser, FormParser])
    def importar_excel(self, request):
        try:
            if 'file' not in request.FILES:
                return Response(
                    {'error': 'No se ha proporcionado ningún archivo'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            excel_file = request.FILES['file']
            allowed_extensions = ['.xlsx', '.xls', '.csv']
            file_extension = os.path.splitext(excel_file.name)[1].lower()

            if file_extension not in allowed_extensions:
                return Response(
                    {'error': 'Formato de archivo no válido. Use Excel o CSV'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Leer el archivo
            if file_extension == '.csv':
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file)

            # Validar columnas requeridas (formato del template)
            headers = [
                'cliente',  # nombre del cliente
                'cliente_documento',
                'tipo_documento',  # dni o ruc
                'fecha_emision',
                'sku',  # SKUs separados por comas
                'productos',  # nombres de productos separados por comas (opcional)
                'cantidades', # separadas por comas
                'precios_unitarios', # separados por comas
                'metodo_pago',
                'igv_incluido',
                'estado',
                'referencia'
            ]
            missing_columns = [col for col in headers if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Faltan columnas requeridas: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            ventas_creadas = []
            errores = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Obtener empresa del usuario
                        empresa = request.user.empresa

                        # Obtener o crear cliente
                        from .models import Cliente
                        tipo_documento = str(row['tipo_documento']).strip().lower()
                        if tipo_documento not in ['dni', 'ruc']:
                            errores.append(f'Error en fila {index + 2}: tipo_documento debe ser "dni" o "ruc"')
                            continue
                        cliente, _ = Cliente.objects.get_or_create(
                            empresa=empresa,
                            documento=str(row['cliente_documento']).strip(),
                            defaults={
                                'nombre': row['cliente'],
                                'tipo_documento': tipo_documento,
                                'direccion': '',
                                'telefono': '',
                                'email': '',
                                'activo': True
                            }
                        )

                        # Procesar SKUs, cantidades y precios_unitarios
                        skus = [p.strip() for p in str(row['sku']).split(',')]
                        cantidades = [c.strip() for c in str(row['cantidades']).split(',')]
                        precios_unitarios = [p.strip().replace('$', '').replace('S/', '').replace('s/', '') for p in str(row['precios_unitarios']).split(',')]

                        if not (len(skus) == len(cantidades) == len(precios_unitarios)):
                            errores.append(f'Error en fila {index + 2}: sku, cantidades y precios_unitarios deben tener la misma cantidad de elementos')
                            continue

                        # Determinar si el IGV está incluido
                        igv_incluido_val = str(row['igv_incluido']).strip().lower()
                        igv_incluido = igv_incluido_val in ['si', 'sí', 's', 'yes', 'y']

                        # Crear la venta
                        from .models import Venta, DetalleVenta
                        venta = Venta.objects.create(
                            empresa=empresa,
                            cliente=cliente,
                            fecha_emision=pd.to_datetime(row['fecha_emision']).date(),
                            estado=row['estado'],
                            tipo_venta='contado',  # Por defecto contado, se puede ajustar según el archivo
                            metodo_pago=row['metodo_pago'],
                            igv_incluido=igv_incluido,
                            referencia=row.get('referencia', '')
                        )

                        # Crear los detalles de venta
                        from apps.inventario.models import Producto
                        for i in range(len(skus)):
                            sku = skus[i]
                            cantidad = Decimal(str(cantidades[i]).replace(',', ''))
                            precio_unitario = Decimal(str(precios_unitarios[i]).replace(',', ''))
                            if not igv_incluido:
                                precio_unitario = precio_unitario * Decimal('1.18')
                            producto, _ = Producto.objects.get_or_create(
                                sku=sku,
                                empresa=empresa,
                                defaults={'nombre': sku}
                            )
                            DetalleVenta.objects.create(
                                venta=venta,
                                producto=producto,
                                cantidad=cantidad,
                                precio_unitario=precio_unitario
                            )

                        # Actualizar los totales de la venta
                        venta.actualizar_totales()
                        ventas_creadas.append(venta.id)

                    except Exception as e:
                        errores.append(f'Error en fila {index + 2}: {str(e)}')

            return Response({
                'mensaje': f'Se importaron {len(ventas_creadas)} ventas correctamente',
                'ventas_creadas': ventas_creadas,
                'errores': errores
            })

        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def estadisticas_ultimos_30_dias(self, request):
        """
        Retorna el número de ventas en los últimos 30 días
        """
        try:
            fecha_inicio = timezone.now() - timezone.timedelta(days=30)
            ventas = self.get_queryset().filter(
                fecha_emision__gte=fecha_inicio,
                estado='pagado'
            ).count()
            
            return Response({
                'ventas_ultimos_30_dias': ventas
            })
        except Exception as e:
            print(f"Error al obtener estadísticas: {str(e)}")
            return Response(
                {'error': f'Error al obtener estadísticas: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def total_ultimos_30_dias(self, request):
        """
        Retorna el total de ventas en los últimos 30 días
        """
        try:
            fecha_inicio = timezone.now() - timezone.timedelta(days=30)
            total = self.get_queryset().filter(
                fecha_emision__gte=fecha_inicio,
                estado='pagado'
            ).aggregate(
                total=Sum('total')
            )['total'] or Decimal('0')
            
            return Response({
                'total_ventas_ultimos_30_dias': float(total)
            })
        except Exception as e:
            print(f"Error al calcular total: {str(e)}")
            return Response(
                {'error': f'Error al calcular total: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def igv_recaudado(self, request):
        """
        Retorna el IGV total recaudado en el mes actual
        """
        try:
            hoy = timezone.now()
            primer_dia_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            igv_total = self.get_queryset().filter(
                fecha_emision__gte=primer_dia_mes,
                estado='pagado'
            ).aggregate(
                total_igv=Sum('igv')
            )['total_igv'] or Decimal('0')
            
            return Response({
                'igv_recaudado': float(igv_total)
            })
        except Exception as e:
            print(f"Error al calcular IGV: {str(e)}")
            return Response(
                {'error': f'Error al calcular IGV: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def productos_mas_vendidos(self, request):
        empresa = request.user.empresa
        productos = (
            DetalleVenta.objects
            .filter(venta__empresa=empresa, venta__estado='pagado')
            .values('producto__nombre')
            .annotate(total_vendido=Sum('cantidad'))
            .order_by('-total_vendido')[:10]
        )
        labels = [p['producto__nombre'] for p in productos]
        data = [int(p['total_vendido']) for p in productos]
        return Response({'labels': labels, 'data': data})

    @action(detail=False, methods=['get'])
    def mejores_clientes(self, request):
        empresa = request.user.empresa
        # Ranking por facturación
        ranking_facturacion = (
            Venta.objects
            .filter(empresa=empresa, estado='pagado')
            .values('cliente__id', 'cliente__nombre', 'cliente__documento', 'cliente__email')
            .annotate(total_facturado=Sum('total'), cantidad_compras=models.Count('id'))
            .order_by('-total_facturado')[:10]
        )
        # Ranking por cantidad de compras
        ranking_recurrencia = (
            Venta.objects
            .filter(empresa=empresa, estado='pagado')
            .values('cliente__id', 'cliente__nombre', 'cliente__documento', 'cliente__email')
            .annotate(total_facturado=Sum('total'), cantidad_compras=models.Count('id'))
            .order_by('-cantidad_compras')[:10]
        )
        return Response({
            'ranking_facturacion': list(ranking_facturacion),
            'ranking_recurrencia': list(ranking_recurrencia)
        })

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        try:
            venta = self.get_object()
            
            # Verificar si la venta puede ser eliminada
            if venta.estado == 'pagada':
                return Response(
                    {'error': 'No se puede eliminar una venta pagada'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Si la venta tiene detalles, eliminarlos primero
            venta.detalles.all().delete()
            
            # Si la venta tiene comprobante, eliminarlo
            if hasattr(venta, 'comprobante_pago'):
                venta.comprobante_pago.delete()
            
            # Eliminar la venta
            venta.delete()
            
            return Response(
                {'mensaje': 'Venta eliminada correctamente'},
                status=status.HTTP_204_NO_CONTENT
            )
        except Exception as e:
            print(f"Error al eliminar venta: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def descargar_template(self, request):
        """
        Descarga el template de importación de ventas en Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Ventas"

        headers = [
            'cliente',  # nombre del cliente
            'cliente_documento',
            'tipo_documento',  # dni o ruc
            'fecha_emision',
            'sku',  # SKUs separados por comas
            'productos',  # nombres de productos separados por comas (opcional)
            'cantidades', # separadas por comas
            'precios_unitarios', # separados por comas
            'metodo_pago',
            'igv_incluido',
            'estado',
            'referencia'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=template_ventas.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Endpoint para obtener estadísticas de cuentas por cobrar
        """
        try:
            empresa = request.user.empresa
            ventas = self.get_queryset().filter(
                tipo_venta__in=['credito_30', 'credito_60'],
                estado='pendiente'
            )
            
            # Calcular total por cobrar y ventas pendientes
            total_por_cobrar = ventas.aggregate(
                total=Sum(F('total') - F('pagos_total'))
            )['total'] or 0
            
            ventas_pendientes = ventas.count()
            
            return Response({
                'total_por_cobrar': total_por_cobrar,
                'ventas_pendientes': ventas_pendientes
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get', 'post'])
    def pagos(self, request, pk=None):
        """
        GET: Obtener los pagos de una venta
        POST: Registrar un nuevo pago para una venta
        """
        try:
            venta = self.get_object()
            print(f"Procesando pago para venta {venta.id}")
            print(f"Datos recibidos: {request.data}")
            
            if request.method == 'GET':
                pagos = PagoVenta.objects.filter(venta=venta)
                serializer = PagoVentaSerializer(pagos, many=True)
                return Response(serializer.data)
            
            elif request.method == 'POST':
                # Verificar que sea una venta a crédito
                if venta.tipo_venta not in ['credito_30', 'credito_60']:
                    return Response(
                        {'error': f'Esta venta no es a crédito. Tipo actual: {venta.tipo_venta}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Verificar que la venta esté pendiente
                if venta.estado != 'pendiente':
                    return Response(
                        {'error': f'Solo se pueden registrar pagos para ventas pendientes. Estado actual: {venta.estado}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Obtener pagos existentes y calcular saldo pendiente
                pagos = PagoVenta.objects.filter(venta=venta).exclude(metodo_pago='pendiente')
                total_pagado = pagos.aggregate(total=Sum('monto'))['total'] or 0
                saldo_pendiente = venta.total - total_pagado
                
                print(f"Total pagado: {total_pagado}")
                print(f"Saldo pendiente: {saldo_pendiente}")
                
                # Validar que el monto del nuevo pago no exceda el saldo pendiente
                try:
                    monto_nuevo_pago = Decimal(str(request.data.get('monto', '0')))
                    print(f"Monto nuevo pago: {monto_nuevo_pago}")
                except (TypeError, ValueError, InvalidOperation) as e:
                    return Response(
                        {'error': f'Monto inválido: {request.data.get("monto")}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                if monto_nuevo_pago <= 0:
                    return Response(
                        {'error': 'El monto del pago debe ser mayor a 0'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                if monto_nuevo_pago > saldo_pendiente:
                    return Response(
                        {'error': f'El monto del pago (S/ {monto_nuevo_pago}) no puede ser mayor al saldo pendiente (S/ {saldo_pendiente})'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Crear el nuevo pago
                data = request.data.copy()
                data['venta'] = venta.id
                serializer = PagoVentaSerializer(data=data)
                
                if serializer.is_valid():
                    print(f"Datos validados: {serializer.validated_data}")
                    pago = serializer.save()
                    
                    # Verificar si la venta está completamente pagada
                    nuevo_total_pagado = total_pagado + monto_nuevo_pago
                    if nuevo_total_pagado >= venta.total:
                        venta.estado = 'pagado'
                        venta.save()
                    
                    return Response({
                        'pago': PagoVentaSerializer(pago).data,
                        'venta_completamente_pagada': nuevo_total_pagado >= venta.total,
                        'saldo_restante': float(venta.total - nuevo_total_pagado)
                    }, status=status.HTTP_201_CREATED)
                
                print(f"Errores de validación: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            print(f"Error al procesar pago: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'])
    def saldo(self, request, pk=None):
        """
        Obtener el saldo pendiente de una venta
        """
        venta = self.get_object()
        pagos = PagoVenta.objects.filter(venta=venta).exclude(metodo_pago='pendiente')
        total_pagado = pagos.aggregate(total=Sum('monto'))['total'] or 0
        saldo_pendiente = venta.total - total_pagado
        
        return Response({
            'total_venta': float(venta.total),
            'total_pagado': float(total_pagado),
            'saldo_pendiente': float(saldo_pendiente)
        })

class PagoVentaViewSet(viewsets.ModelViewSet):
    queryset = PagoVenta.objects.all()
    serializer_class = PagoVentaSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        queryset = PagoVenta.objects.filter(venta__empresa=self.request.user.empresa)
        
        # Filtrar por venta específica si se proporciona el ID
        venta_id = self.kwargs.get('venta_pk')
        if venta_id:
            queryset = queryset.filter(venta_id=venta_id)
        
        return queryset.select_related('venta', 'venta__cliente')

    def perform_create(self, serializer):
        venta_id = self.kwargs.get('venta_pk')
        if venta_id:
            venta = get_object_or_404(Venta, pk=venta_id, empresa=self.request.user.empresa)
            
            # Verificar que la venta sea a crédito
            if venta.tipo_venta not in ['credito_30', 'credito_60']:
                raise serializers.ValidationError('Solo se pueden registrar pagos para ventas a crédito')
            
            # Verificar que la venta esté pendiente
            if venta.estado != 'pendiente':
                raise serializers.ValidationError('Solo se pueden registrar pagos para ventas pendientes')
            
            # Obtener el saldo pendiente
            pagos_existentes = PagoVenta.objects.filter(venta=venta).exclude(metodo_pago='pendiente')
            total_pagado = pagos_existentes.aggregate(total=Sum('monto'))['total'] or 0
            saldo_pendiente = venta.total - total_pagado
            
            # Verificar que el monto del nuevo pago no exceda el saldo pendiente
            monto_nuevo_pago = Decimal(str(self.request.data.get('monto', 0)))
            if monto_nuevo_pago > saldo_pendiente:
                raise serializers.ValidationError(
                    f'El monto del pago (S/ {monto_nuevo_pago}) no puede ser mayor al saldo pendiente (S/ {saldo_pendiente})'
                )
            
            serializer.save(venta=venta)
            
            # Actualizar estado de la venta si está completamente pagada
            nuevo_total_pagado = total_pagado + monto_nuevo_pago
            if nuevo_total_pagado >= venta.total:
                venta.estado = 'pagado'
                venta.save()
        else:
            raise serializers.ValidationError('Se requiere el ID de la venta')

    def perform_update(self, serializer):
        pago_anterior = self.get_object()
        monto_anterior = pago_anterior.monto
        venta = pago_anterior.venta
        
        # Verificar que la venta sea a crédito
        if venta.tipo_venta not in ['credito_30', 'credito_60']:
            raise serializers.ValidationError('Solo se pueden modificar pagos de ventas a crédito')
        
        # Calcular saldo pendiente sin contar el pago actual
        pagos_existentes = PagoVenta.objects.filter(venta=venta).exclude(id=pago_anterior.id).exclude(metodo_pago='pendiente')
        total_pagado = pagos_existentes.aggregate(total=Sum('monto'))['total'] or 0
        saldo_pendiente = venta.total - total_pagado
        
        # Verificar que el nuevo monto no exceda el saldo pendiente
        monto_nuevo = Decimal(str(self.request.data.get('monto', 0)))
        if monto_nuevo > saldo_pendiente:
            raise serializers.ValidationError(
                f'El monto del pago (S/ {monto_nuevo}) no puede ser mayor al saldo pendiente (S/ {saldo_pendiente})'
            )
        
        serializer.save()
        
        # Actualizar estado de la venta
        nuevo_total_pagado = total_pagado + monto_nuevo
        if nuevo_total_pagado >= venta.total:
            venta.estado = 'pagado'
        else:
            venta.estado = 'pendiente'
        venta.save()

    @action(detail=False, methods=['get'])
    def ventas_pendientes(self, request):
        """
        Obtener todas las ventas a crédito que están pendientes de pago
        """
        ventas = self.get_queryset().filter(
            tipo_venta__in=['credito_30', 'credito_60'],
            estado='pendiente'
        ).annotate(
            total_pagado=CoalesceFunc(
                Sum('pagos__monto', filter=~Q(pagos__metodo_pago='pendiente')),
                0,
                output_field=models.DecimalField()
            ),
            saldo_pendiente=F('total') - F('total_pagado')
        ).filter(saldo_pendiente__gt=0)

        serializer = self.get_serializer(ventas, many=True)
        return Response(serializer.data) 