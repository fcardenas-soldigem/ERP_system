import logging
from django.core.cache import cache
from rest_framework import viewsets, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db.models import Sum, Count, Q, F, ExpressionWrapper, IntegerField
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from apps.core.permissions import HasEmpresaPermission
from .models import Compra, CompraDetalle, Proveedor, OrdenCompra, RecepcionCompra, PagoCompra
from apps.inventario.models import Producto, Almacen

logger = logging.getLogger(__name__)
from .serializers import (
    CompraSerializer,
    CompraDetalleSerializer,
    ProveedorSerializer,
    OrdenCompraSerializer,
    RecepcionCompraSerializer,
    PagoCompraSerializer
)
from rest_framework.views import APIView
import os
from io import BytesIO
from rest_framework.renderers import BaseRenderer
from rest_framework.pagination import PageNumberPagination
from rest_framework.exceptions import PermissionDenied
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
import json

class ExcelRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None
    render_style = 'binary'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data

class CompraPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class CompraViewSet(viewsets.ModelViewSet):
    queryset = Compra.objects.all()
    serializer_class = CompraSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]
    pagination_class = CompraPagination
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        try:
            if not hasattr(self.request.user, 'empresa'):
                raise PermissionDenied('Usuario no tiene empresa asignada')

            empresa = self.request.user.empresa

            queryset = self.queryset.filter(empresa=empresa)
            
            # Filtros adicionales por parámetros de URL
            # Filtro por tipo de compra
            tipo_compra_in = self.request.query_params.get('tipo_compra__in')
            if tipo_compra_in:
                tipos = tipo_compra_in.split(',')
                queryset = queryset.filter(tipo_compra__in=tipos)
            
            # Filtro por estado
            estado_in = self.request.query_params.get('estado__in')
            if estado_in:
                estados = estado_in.split(',')
                queryset = queryset.filter(estado__in=estados)
            
            queryset = queryset.select_related(
                'empresa',
                'proveedor',
                'almacen'
            ).prefetch_related(
                'detalles',
                'detalles__producto',
                'pagos'
            ).order_by('-fecha_emision', '-id')

            return queryset
        except Exception as e:
            logger.error('Error en get_queryset: %s', e, exc_info=True)
            raise

    def retrieve(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            serializer = self.get_serializer(instance)
            return Response(serializer.data)
        except Exception as e:
            logger.error('Error en retrieve: %s', e, exc_info=True)
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    def perform_create(self, serializer):
        if not hasattr(self.request.user, 'empresa'):
            raise PermissionDenied('Usuario no tiene empresa asignada')
        serializer.save(empresa=self.request.user.empresa)

    def list(self, request, *args, **kwargs):
        try:
            empresa_id = request.user.empresa_id
            if not request.query_params:
                cache_key = f'compras_list_{empresa_id}'
                cached = cache.get(cache_key)
                if cached is not None:
                    return Response(cached)

            queryset = self.get_queryset()

            # Aplicar paginación
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                response = self.get_paginated_response(serializer.data)
                if not request.query_params:
                    cache.set(f'compras_list_{empresa_id}', response.data, 60)
                return response

            serializer = self.get_serializer(queryset, many=True)
            data = {
                'count': queryset.count(),
                'next': None,
                'previous': None,
                'results': serializer.data
            }
            if not request.query_params:
                cache.set(f'compras_list_{empresa_id}', data, 60)
            return Response(data)
        except Exception as e:
            logger.error('Error en list: %s', e, exc_info=True)
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    def update(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                instance = self.get_object()
                
                # Verificar permisos
                if instance.empresa != request.user.empresa:
                    return Response(
                        {"detail": "No tienes permiso para actualizar esta compra."},
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                serializer = self.get_serializer(instance, data=request.data, partial=True)
                serializer.is_valid(raise_exception=True)
                self.perform_update(serializer)
                
                return Response(serializer.data)
        except Exception as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def cambiar_estado(self, request, pk=None):
        compra = self.get_object()
        nuevo_estado = request.data.get('estado')
        
        if nuevo_estado not in ['BORRADOR', 'CONFIRMADO', 'ANULADO']:
            return Response(
                {"detail": "Estado no válido"},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        compra.estado = nuevo_estado
        compra.save()
        
        serializer = self.get_serializer(compra)
        return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='bulk-estado')
    def bulk_cambiar_estado(self, request):
        """
        Cambia el estado de múltiples compras en una sola operación.
        Solo permite transiciones válidas; las anuladas no se pueden reactivar.
        Seguridad multi-tenant: filtra por empresa del usuario autenticado.
        """
        ids = request.data.get('ids', [])
        nuevo_estado = request.data.get('estado', '').lower()

        VALID_STATES = {'borrador', 'pendiente', 'pagada', 'anulada'}
        VALID_TRANSITIONS = {
            'borrador':  {'pendiente', 'pagada', 'anulada'},
            'pendiente': {'pagada', 'anulada'},
            'pagada':    {'anulada'},
            'anulada':   set(),
        }

        if not ids or not isinstance(ids, list):
            return Response({'detail': 'El campo ids es requerido y debe ser una lista.'},
                            status=status.HTTP_400_BAD_REQUEST)
        if nuevo_estado not in VALID_STATES:
            return Response({'detail': f'Estado "{nuevo_estado}" no válido. Opciones: {", ".join(sorted(VALID_STATES))}'},
                            status=status.HTTP_400_BAD_REQUEST)

        empresa = request.user.empresa
        compras_qs = Compra.objects.filter(id__in=ids, empresa=empresa)
        encontradas = {c.id for c in compras_qs}

        errores = []
        for id_solicitado in ids:
            if id_solicitado not in encontradas:
                errores.append({'id': id_solicitado, 'error': 'No encontrada o no pertenece a tu empresa.'})

        # Validar transiciones individualmente
        actualizables = []
        for compra in compras_qs:
            permitidos = VALID_TRANSITIONS.get(compra.estado, set())
            if nuevo_estado in permitidos:
                actualizables.append(compra)
            else:
                errores.append({
                    'id': compra.id,
                    'numero': compra.numero,
                    'error': f'Transición {compra.estado} → {nuevo_estado} no permitida.'
                })

        if not actualizables and errores:
            return Response({'detail': 'Ninguna compra pudo ser actualizada.', 'errores': errores},
                            status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for compra in actualizables:
                compra.estado = nuevo_estado
                compra.save(update_fields=['estado', 'updated_at'])

        serializer = self.get_serializer(actualizables, many=True)
        return Response({
            'actualizadas': len(actualizables),
            'errores': errores,
            'compras': serializer.data,
        })

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        compras = self.get_queryset()
        from io import BytesIO
        import pandas as pd
        from openpyxl.utils import get_column_letter
        from django.utils import timezone
        from django.http import HttpResponse

        datos = []
        for compra in compras:
            for detalle in compra.detalles.all():
                datos.append({
                    'Número': compra.numero,
                    'Fecha': compra.fecha_emision.strftime('%Y-%m-%d') if compra.fecha_emision else '',
                    'Proveedor': str(compra.proveedor) if compra.proveedor else '',
                    'Producto': detalle.producto.nombre if detalle.producto else '',
                    'Cantidad': float(detalle.cantidad),
                    'Precio Unitario': float(detalle.precio_unitario),
                    'IGV Incluido': 'Sí' if compra.igv_incluido else 'No',
                    'Subtotal': float(compra.subtotal),
                    'IGV': float(compra.igv),
                    'Total': float(compra.total),
                    'Estado': compra.get_estado_display(),
                    'Método de Pago': compra.get_metodo_pago_display(),
                    'Tipo de Compra': compra.get_tipo_compra_display(),
                    'Saldo Pendiente': float(compra.get_saldo_pendiente())
                })

        df = pd.DataFrame(datos)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Compras', index=False)
            worksheet = writer.sheets['Compras']
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(str(col))
                ) + 2
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = min(max_length, 50)
        output.seek(0)
        filename = f'Compras_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    @action(detail=False, methods=['POST'])
    def importar_excel(self, request):
        try:
            if 'file' not in request.FILES:
                return Response(
                    {'error': 'No se ha proporcionado ningún archivo'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            excel_file = request.FILES['file']
            allowed_extensions = ['.xlsx', '.xls', '.csv']
            file_extension = os.path.splitext(excel_file.name)[1].lower()

            if file_extension not in allowed_extensions:
                return Response(
                    {'error': 'Formato de archivo no válido. Use Excel o CSV'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Leer el archivo
            if file_extension == '.csv':
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file)
            
            # Validar columnas requeridas (ajustadas a tu template)
            required_columns = ['empresa', 'proveedor', 'almacen', 'fecha', 'sku', 'nombre_producto', 'cantidad', 'precio_unitario', 'subtotal', 'igv_incluido', 'estado', 'metodo_pago']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Faltan columnas requeridas: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
                
            compras_creadas = []
            errores = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        # Obtener empresa
                        from apps.empresas.models import Empresa
                        empresa = Empresa.objects.get(nombre=row['empresa'])

                        # Obtener o crear proveedor
                        proveedor = Proveedor.objects.get(
                            razon_social=row['proveedor'],
                            empresa=empresa
                        )

                        # Obtener almacén
                        almacen = Almacen.objects.get(
                            nombre=row['almacen'],
                            empresa=empresa
                        )

                        # Obtener o crear producto
                        from apps.inventario.models import Producto
                        producto, creado = Producto.objects.get_or_create(
                            sku=str(row['sku']),
                            empresa=empresa,
                            defaults={
                                'nombre': row.get('nombre_producto', row['sku'])
                            }
                        )

                        # Procesar cantidad y precio_unitario (eliminar comas si las hay)
                        cantidad = Decimal(str(row['cantidad']).replace(',', ''))
                        precio_unitario = Decimal(str(row['precio_unitario']).replace(',', '').replace('$', ''))
                    
                        # Crear la compra
                        compra = Compra.objects.create(
                            empresa=empresa,
                            proveedor=proveedor,
                            almacen=almacen,
                            fecha_emision=pd.to_datetime(row['fecha']).date(),
                            estado=row['estado'],
                            metodo_pago=row['metodo_pago'],
                            igv_incluido=str(row['igv_incluido']).lower() == 'no'  # Si dice "no", significa que NO está incluido
                        )

                        # Crear el detalle de compra
                        CompraDetalle.objects.create(
                            compra=compra,
                            producto=producto,
                            cantidad=cantidad,
                            precio_unitario=precio_unitario
                        )

                        # Actualizar los totales de la compra
                        compra.actualizar_totales()
                        compras_creadas.append(compra.id)

                    except Exception as e:
                        errores.append(f'Error en fila {index + 2}: {str(e)}')
                    
            return Response({
                'mensaje': f'Se importaron {len(compras_creadas)} compras correctamente',
                'compras_creadas': compras_creadas,
                'errores': errores
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """
        Endpoint para obtener estadísticas de cuentas por pagar
        """
        try:
            empresa = request.user.empresa
            compras = self.get_queryset().filter(
                tipo_compra__in=['credito_30', 'credito_60'],
                estado__in=['pendiente', 'borrador']
            ).exclude(estado='anulada')
            
            # Calcular total por pagar y compras pendientes
            total_por_pagar = sum(compra.get_saldo_pendiente() for compra in compras)
            compras_pendientes = compras.count()
            
            return Response({
                'total_por_pagar': total_por_pagar,
                'compras_pendientes': compras_pendientes
            })
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get', 'post'])
    def pagos(self, request, pk=None):
        try:
            compra = self.get_object()
            
            if request.method == 'GET':
                # Obtener los pagos existentes
                pagos = PagoCompra.objects.filter(compra=compra)
                serializer = PagoCompraSerializer(pagos, many=True)
                return Response(serializer.data)
            
            elif request.method == 'POST':
                # Crear un nuevo pago
                
                # Validar que la compra sea a crédito o tenga saldo pendiente
                if compra.tipo_compra not in ['credito_30', 'credito_60'] and compra.estado == 'pagada':
                    return Response(
                        {'error': 'No se pueden registrar pagos para esta compra'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Obtener el saldo pendiente
                saldo_pendiente = compra.get_saldo_pendiente()
                if saldo_pendiente <= 0:
                    return Response(
                        {'error': 'Esta compra no tiene saldo pendiente'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Validar el monto del nuevo pago
                try:
                    monto_nuevo_pago = Decimal(str(request.data.get('monto', '0')))
                except (TypeError, ValueError) as e:
                    return Response(
                        {'error': f'Monto inválido: {request.data.get("monto")}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                if monto_nuevo_pago <= 0:
                    return Response(
                        {'error': 'El monto del pago debe ser mayor a 0'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                if monto_nuevo_pago > saldo_pendiente:
                    return Response(
                        {'error': f'El monto del pago (S/ {monto_nuevo_pago}) no puede ser mayor al saldo pendiente (S/ {saldo_pendiente})'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Crear el nuevo pago
                data = request.data.copy()
                data['compra'] = compra.id
                serializer = PagoCompraSerializer(data=data)
                
                if serializer.is_valid():
                    pago = serializer.save()
                    
                    # Actualizar el estado de la compra si es necesario
                    pagos_total = compra.pagos.aggregate(total=Sum('monto'))['total'] or 0
                    if pagos_total >= compra.total:
                        compra.estado = 'pagada'
                        compra.save()
                    
                    return Response({
                        'pago': PagoCompraSerializer(pago).data,
                        'compra_completamente_pagada': pagos_total >= compra.total,
                        'saldo_restante': float(compra.total - pagos_total)
                    }, status=status.HTTP_201_CREATED)
                
                logger.debug('Errores de validación pago compra: %s', serializer.errors)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            logger.error('Error en pagos: %s', e, exc_info=True)
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def compras_pendientes(self, request):
        """
        Endpoint para obtener las compras pendientes de pago con conversión de moneda
        """
        try:
            from apps.core.services.documento_service import DocumentoService
            
            empresa = request.user.empresa
            compras = self.get_queryset().filter(
                empresa=empresa,
                tipo_compra__in=['credito_30', 'credito_60'],
                estado__in=['pendiente', 'borrador']
            ).exclude(
                estado='anulada'
            )
            
            # Obtener tipo de cambio del día
            doc_service = DocumentoService()
            try:
                tipo_cambio_data = doc_service.obtener_tipo_cambio_venta()
                tipo_cambio = float(tipo_cambio_data.get('venta', 3.8)) if tipo_cambio_data else 3.8
            except:
                tipo_cambio = 3.8  # Fallback
            
            # Procesar compras con conversión de moneda
            compras_procesadas = []
            total_general_pen = 0
            
            for compra in compras:
                # Calcular saldo pendiente
                pagos_reales = compra.pagos.exclude(metodo_pago='pendiente')
                total_pagado = sum([float(pago.monto) for pago in pagos_reales])
                saldo_pendiente = float(compra.total) - total_pagado
                
                if saldo_pendiente > 0:  # Solo incluir si tiene saldo pendiente
                    # Convertir a PEN si la compra es en USD
                    saldo_pendiente_pen = saldo_pendiente
                    total_pen = float(compra.total)
                    
                    if compra.moneda == 'USD':
                        saldo_pendiente_pen = saldo_pendiente * tipo_cambio
                        total_pen = float(compra.total) * tipo_cambio
                    
                    # Calcular días restantes hasta vencimiento
                    dias_restantes = None
                    if compra.fecha_vencimiento:
                        dias_restantes = (compra.fecha_vencimiento - timezone.now().date()).days
                    
                    compra_data = {
                        'id': compra.id,
                        'numero': compra.numero,
                        'proveedor': {
                            'id': compra.proveedor.id,
                            'razon_social': compra.proveedor.razon_social,
                            'ruc': compra.proveedor.ruc
                        },
                        'fecha_emision': compra.fecha_emision,
                        'fecha_vencimiento': compra.fecha_vencimiento,
                        'tipo_compra': compra.tipo_compra,
                        'estado': compra.estado,
                        'moneda': compra.moneda,
                        'total': float(compra.total),
                        'total_pen': total_pen,  # Total convertido a PEN
                        'saldo_pendiente': saldo_pendiente,
                        'saldo_pendiente_pen': saldo_pendiente_pen,  # Saldo convertido a PEN
                        'dias_restantes': dias_restantes
                    }
                    
                    compras_procesadas.append(compra_data)
                    total_general_pen += saldo_pendiente_pen
            
            return Response({
                'compras': compras_procesadas,
                'total_general_pen': total_general_pen,
                'tipo_cambio': tipo_cambio,
                'count': len(compras_procesadas)
            })
        except Exception as e:
            logger.error('Error en compras_pendientes: %s', e, exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'], url_path='saldo-pendiente')
    def saldo_pendiente(self, request, pk=None):
        try:
            compra = self.get_object()
            saldo = compra.get_saldo_pendiente()
            return Response({'saldo_pendiente': saldo})
        except Exception as e:
            logger.error('Error en saldo_pendiente: %s', e, exc_info=True)
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request, pk=None):
        """
        Exportar Orden de Compra a PDF profesional
        """
        import traceback
        compra = self.get_object()
        
        try:
            from .utils.pdf_generator import OrdenCompraPDFGenerator
            
            pdf_generator = OrdenCompraPDFGenerator(compra)
            pdf_buffer = pdf_generator.generar_pdf()
            
            # Preparar respuesta
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            filename = f'OC_{compra.numero}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
        
        except Exception as e:
            error_traceback = traceback.format_exc()
            logger.error("Error al generar PDF: %s", e, exc_info=True)
            return Response(
                {'error': f'Error al generar PDF: {str(e)}', 'detail': error_traceback},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['GET'])
    def get_productos(self, request):
        """
        Endpoint para obtener los productos disponibles para compras
        """
        try:
            from apps.inventario.models import Producto
            from apps.inventario.serializers import ProductoSerializer
            
            # Asegurarnos de obtener la empresa correcta
            empresa = request.user.empresa
            if not empresa:
                return Response(
                    {'error': 'No se encontró una empresa asociada al usuario'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            productos = Producto.objects.filter(
                empresa=empresa,
                is_active=True
            ).order_by('nombre')
            
            serializer = ProductoSerializer(productos, many=True)
            return Response(serializer.data)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    def create(self, request, *args, **kwargs):
        """
        Método personalizado para crear compras con detalles de productos
        """
        try:
            
            with transaction.atomic():
                # Obtener datos básicos de la compra
                empresa = request.user.empresa
                if not empresa:
                    return Response(
                        {'error': 'Usuario no tiene empresa asignada'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Procesar detalles si existen
                detalles_data = []
                if 'detalles' in request.data:
                    try:
                        detalles_json = request.data.get('detalles')
                        if isinstance(detalles_json, str):
                            detalles_data = json.loads(detalles_json)
                        else:
                            detalles_data = detalles_json
                    except Exception as e:
                        logger.error('Error al procesar detalles: %s', e, exc_info=True)
                        return Response(
                            {'error': f'Error al procesar detalles: {str(e)}'},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                # Validar que haya detalles
                if not detalles_data:
                    return Response(
                        {'error': 'Debe incluir al menos un producto en la compra'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Crear datos para el serializer (sin detalles)
                compra_data = request.data.copy()
                if 'detalles' in compra_data:
                    del compra_data['detalles']

                # Crear la compra usando el serializer
                serializer = self.get_serializer(data=compra_data)
                serializer.is_valid(raise_exception=True)
                compra = serializer.save(empresa=empresa)


                # Crear los detalles
                from .models import CompraDetalle
                from apps.inventario.models.producto import Producto
                
                
                for idx, detalle_data in enumerate(detalles_data):
                    try:
                        
                        # Validar datos del detalle
                        producto_id = detalle_data.get('producto')
                        cantidad = detalle_data.get('cantidad')
                        precio_unitario = detalle_data.get('precio_unitario')


                        if not producto_id or not cantidad or not precio_unitario:
                            raise ValueError('Faltan datos en el detalle del producto')

                        # Obtener el producto
                        producto = Producto.objects.get(
                            id=int(producto_id),
                            empresa=empresa
                        )

                        # Crear el detalle
                        detalle = CompraDetalle.objects.create(
                            compra=compra,
                            producto=producto,
                            cantidad=Decimal(str(cantidad)),
                            precio_unitario=Decimal(str(precio_unitario))
                        )

                    except Exception as e:
                        logger.error('Error al crear detalle: %s', e, exc_info=True)
                        raise ValueError(f'Error al procesar producto: {str(e)}')
                

                # Actualizar totales de la compra
                compra.actualizar_totales()

                # Si la compra está pagada, actualizar el stock inmediatamente
                if compra.estado == 'pagada':
                    compra.actualizar_stock()

                # Retornar la compra creada
                empresa_id = request.user.empresa_id
                cache.delete(f'compras_list_{empresa_id}')
                cache.delete(f'dashboard_mes_{empresa_id}')
                cache.delete(f'dashboard_historico_{empresa_id}')
                cache.delete(f'dashboard_simple_stats_{empresa_id}')
                response_serializer = self.get_serializer(compra)
                return Response(response_serializer.data, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error('Error en create CompraViewSet: %s', e, exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        return self.queryset.filter(empresa=self.request.user.empresa)

    def list(self, request, *args, **kwargs):
        empresa_id = request.user.empresa_id
        cache_key = f'proveedores_list_{empresa_id}'
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)
        response = super().list(request, *args, **kwargs)
        cache.set(cache_key, response.data, 300)
        return response

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)
        cache.delete(f'proveedores_list_{self.request.user.empresa_id}')

    def perform_update(self, serializer):
        serializer.save()
        cache.delete(f'proveedores_list_{self.request.user.empresa_id}')

    def perform_destroy(self, instance):
        instance.delete()
        cache.delete(f'proveedores_list_{self.request.user.empresa_id}')

    @action(detail=False, methods=['get'])
    def consultar_ruc(self, request):
        """
        Consulta automática de datos de proveedor por RUC usando la API de SUNAT
        """
        ruc = request.query_params.get('ruc')
        
        if not ruc:
            return Response(
                {'error': 'Debe proporcionar un RUC'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(ruc) != 11:
            return Response(
                {'error': 'El RUC debe tener 11 dígitos'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from apps.core.services.documento_service import DocumentoService
            
            # Consultar datos del RUC
            resultado = DocumentoService.consultar_ruc(ruc)
            
            if resultado.get('success'):
                data = resultado['data']
                response_data = {
                    'success': True,
                    'data': {
                        'ruc': data.get('ruc', ruc),
                        'razon_social': data.get('razon_social', ''),
                        'direccion': data.get('direccion', ''),
                        'estado': data.get('estado', ''),
                        'condicion': data.get('condicion', ''),
                        'ubigeo': data.get('ubigeo', ''),
                    }
                }
                return Response(response_data)
            else:
                return Response(
                    {'error': resultado.get('error', 'No se pudo consultar el RUC')}, 
                    status=status.HTTP_404_NOT_FOUND
                )
                
        except Exception as e:
            return Response(
                {'error': f'Error interno del servidor: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(
        detail=False,
        methods=['get'],
        renderer_classes=[ExcelRenderer],
    )
    def exportar(self, request):
        try:
            proveedores = self.get_queryset()
            
            if not proveedores.exists():
                return Response(
                    {'error': 'No hay proveedores para exportar'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Crear el archivo Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Proveedores"

            # Definir las columnas
            columns = [
                'RUC',
                'Razón Social',
                'Teléfono',
                'Estado',
                'Fecha Registro'
            ]

            # Escribir los encabezados
            for col_num, column_title in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = cell.font.copy(bold=True)

            # Escribir los datos
            for row_num, proveedor in enumerate(proveedores, 2):
                row = [
                    proveedor.ruc,
                    proveedor.razon_social,
                    proveedor.telefono or '',
                    'Activo' if proveedor.activo else 'Inactivo',
                    proveedor.created_at.strftime('%Y-%m-%d') if proveedor.created_at else ''
                ]
                
                for col_num, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = cell_value

            # Ajustar el ancho de las columnas
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

            # Guardar el Excel en memoria
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Preparar la respuesta
            response = HttpResponse(
                excel_file.getvalue(),
                content_type=ExcelRenderer.media_type
            )
            response['Content-Disposition'] = f'attachment; filename=Proveedores_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
            
            return response

        except Exception as e:
            logger.error("Error en exportar compras: %s", e, exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class OrdenCompraViewSet(viewsets.ModelViewSet):
    queryset = OrdenCompra.objects.all()
    serializer_class = OrdenCompraSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        qs = self.queryset.filter(empresa=self.request.user.empresa)
        estado = self.request.query_params.get('estado')
        if estado:
            qs = qs.filter(estado=estado)
        return qs

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

    @action(detail=True, methods=['post'], url_path='cambiar-estado')
    def cambiar_estado(self, request, pk=None):
        orden = self.get_object()
        nuevo_estado = request.data.get('estado')
        estados_validos = [s[0] for s in OrdenCompra.ESTADO_CHOICES]
        if nuevo_estado not in estados_validos:
            return Response(
                {'error': f'Estado inválido. Opciones: {estados_validos}'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        orden.estado = nuevo_estado
        orden.save(update_fields=['estado'])
        return Response(OrdenCompraSerializer(orden).data)

    @action(detail=True, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request, pk=None):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from datetime import datetime

        orden = self.get_object()
        empresa = orden.empresa
        buffer = BytesIO()

        PRIMARY = colors.HexColor('#2B6CB0')
        LIGHT = colors.HexColor('#EBF8FF')
        DARK = colors.HexColor('#1a1a1a')
        GRAY = colors.HexColor('#666666')
        BORDER = colors.HexColor('#cccccc')
        WHITE = colors.white
        PAGE_W, _ = A4
        MARGIN = 40

        styles = getSampleStyleSheet()
        lbl = ParagraphStyle('lbl', fontSize=8, textColor=GRAY)
        val = ParagraphStyle('val', fontSize=10, fontName='Helvetica-Bold', textColor=DARK)
        bar_s = ParagraphStyle('bar', fontSize=9, fontName='Helvetica-Bold', textColor=WHITE, leading=12)
        cell_s = ParagraphStyle('cell', fontSize=9, textColor=DARK)
        cell_r = ParagraphStyle('cellr', fontSize=9, textColor=DARK, alignment=TA_RIGHT)

        def section_bar(text):
            t = Table([[Paragraph(text, bar_s)]], colWidths=[PAGE_W - 2 * MARGIN])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), PRIMARY),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ]))
            return t

        def lv(label, value):
            return Table(
                [[Paragraph(label, lbl)], [Paragraph(str(value) if value else '—', val)]],
                colWidths=[(PAGE_W - 2 * MARGIN) / 2],
            )

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=MARGIN, bottomMargin=MARGIN + 10,
        )

        def page_decor(c, d):
            c.saveState()
            c.setFont('Helvetica', 7)
            c.setFillColor(GRAY)
            c.drawString(MARGIN, 20, f'OC-{orden.numero or "---"}')
            c.drawRightString(PAGE_W - MARGIN, 20, datetime.now().strftime('%d/%m/%Y %H:%M'))
            c.drawCentredString(PAGE_W / 2, 20, f'Página {c.getPageNumber()}')
            c.restoreState()

        W = PAGE_W - 2 * MARGIN

        # Header
        right_data = [
            [Paragraph(f'Fecha: {orden.fecha_creacion.strftime("%d/%m/%Y")}',
                       ParagraphStyle('h', fontSize=8, textColor=GRAY, alignment=TA_RIGHT))],
            [Paragraph('ORDEN DE COMPRA', ParagraphStyle('h2', fontSize=10, fontName='Helvetica-Bold',
                                                          textColor=DARK, alignment=TA_RIGHT, spaceBefore=4))],
            [Paragraph(f'<b>OC-{orden.numero or "---"}</b>',
                       ParagraphStyle('h3', fontSize=16, fontName='Helvetica-Bold',
                                      textColor=PRIMARY, alignment=TA_RIGHT))],
        ]
        right_t = Table(right_data, colWidths=[W * 0.45])
        logo_cell = ''
        if empresa.logo:
            try:
                import os as _os
                from reportlab.platypus import Image as _Img
                logo_path = empresa.logo.path
                if _os.path.exists(logo_path):
                    logo_cell = _Img(logo_path, width=130, height=60, kind='proportional')
            except Exception:
                pass
        if not logo_cell:
            logo_cell = Paragraph(f'<b>{empresa.nombre}</b>',
                                  ParagraphStyle('en', fontSize=14, fontName='Helvetica-Bold',
                                                 textColor=PRIMARY))
        header_t = Table([[logo_cell, right_t]], colWidths=[W * 0.55, W * 0.45])
        header_t.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))

        # Info empresa + proveedor
        emp_rows = [
            [lv('Empresa', empresa.nombre), lv('RUC', empresa.ruc or '—')],
            [lv('Dirección', empresa.direccion or '—'), ''],
        ]
        emp_t = Table(emp_rows, colWidths=[W / 2, W / 2])
        emp_t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.3, BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))

        prov_rows = [
            [lv('Proveedor', orden.proveedor_nombre or '—'), lv('Fecha Entrega', orden.fecha_entrega.strftime('%d/%m/%Y'))],
            [lv('Estado', orden.get_estado_display()), ''],
        ]
        prov_t = Table(prov_rows, colWidths=[W / 2, W / 2])
        prov_t.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.3, BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))

        # Totals
        s_lbl = ParagraphStyle('tl', fontSize=9, textColor=DARK)
        s_val = ParagraphStyle('tv', fontSize=9, textColor=DARK, alignment=TA_RIGHT)
        s_lbl_b = ParagraphStyle('tlb', fontSize=10, fontName='Helvetica-Bold', textColor=DARK)
        s_val_b = ParagraphStyle('tvb', fontSize=10, fontName='Helvetica-Bold', textColor=PRIMARY, alignment=TA_RIGHT)
        sym = 'S/'
        tot_rows = [
            [Paragraph('Subtotal:', s_lbl), Paragraph(f'{sym} {float(orden.subtotal):,.2f}', s_val)],
            [Paragraph('IGV (18%):', s_lbl), Paragraph(f'{sym} {float(orden.igv):,.2f}', s_val)],
            [Paragraph('<b>TOTAL:</b>', s_lbl_b), Paragraph(f'<b>{sym} {float(orden.total):,.2f}</b>', s_val_b)],
        ]
        tot_col = W - 180
        tot_t = Table(
            [[Paragraph('', styles['Normal']), Table(tot_rows, colWidths=[90, 90])]],
            colWidths=[tot_col, 180],
        )
        tot_t.setStyle(TableStyle([
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))

        story = [
            header_t, Spacer(1, 8),
            section_bar('EMPRESA EMISORA'), emp_t, Spacer(1, 4),
            section_bar('DATOS DEL PROVEEDOR Y ENTREGA'), prov_t, Spacer(1, 8),
        ]
        if orden.notas:
            story += [section_bar('NOTAS / OBSERVACIONES'),
                      Paragraph(orden.notas, ParagraphStyle('nn', fontSize=9, textColor=DARK)),
                      Spacer(1, 8)]
        story += [section_bar('RESUMEN ECONÓMICO'), Spacer(1, 4), tot_t, Spacer(1, 30)]

        # Signature area
        sig_t = Table(
            [[Paragraph('_______________________', styles['Normal']),
              Paragraph('_______________________', styles['Normal'])],
             [Paragraph('Firma autorizada', ParagraphStyle('sig', fontSize=8, textColor=GRAY)),
              Paragraph('Visto bueno proveedor', ParagraphStyle('sig2', fontSize=8, textColor=GRAY, alignment=TA_RIGHT))]],
            colWidths=[W / 2, W / 2],
        )
        story.append(sig_t)

        doc.build(story, onFirstPage=page_decor, onLaterPages=page_decor)
        buffer.seek(0)

        prov_slug = (orden.proveedor_nombre or 'PROVEEDOR').replace(' ', '_')[:30]
        fecha_str = orden.fecha_creacion.strftime('%Y-%m-%d')
        filename = f'OC-{orden.numero or "---"}_{prov_slug}_{fecha_str}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class RecepcionCompraViewSet(viewsets.ModelViewSet):
    queryset = RecepcionCompra.objects.all()
    serializer_class = RecepcionCompraSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        return self.queryset.filter(empresa=self.request.user.empresa)

    def perform_create(self, serializer):
        serializer.save(empresa=self.request.user.empresa)

class ResumenComprasMensual(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def list(self, request):
        # Obtener el primer día del mes actual
        today = timezone.now()
        first_day = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Calcular el último día del mes
        if today.month == 12:
            last_day = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            last_day = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        
        # Filtrar compras del mes actual
        compras = Compra.objects.filter(
            empresa=request.user.empresa,
            fecha_emision__range=(first_day, last_day)
        )
        
        # Calcular totales
        total_compras = compras.count()
        monto_total = compras.aggregate(total=Sum('total'))['total'] or Decimal('0')
        
        return Response({
            'total_compras': total_compras,
            'monto_total': float(monto_total),
            'mes': today.strftime('%B %Y')
        })

    def retrieve(self, request, pk=None):
        # Obtener el mes y año específico
        try:
            year, month = map(int, pk.split('-'))
            first_day = timezone.datetime(year, month, 1)
            if month == 12:
                last_day = timezone.datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = timezone.datetime(year, month + 1, 1) - timedelta(days=1)
            
            compras = Compra.objects.filter(
                empresa=request.user.empresa,
                fecha_emision__range=(first_day, last_day)
            )
            
            total_compras = compras.count()
            monto_total = compras.aggregate(total=Sum('total'))['total'] or Decimal('0')
            
            return Response({
                'total_compras': total_compras,
                'monto_total': float(monto_total),
                'mes': first_day.strftime('%B %Y')
            })
        except (ValueError, IndexError):
            return Response(
                {'detail': 'Formato de fecha inválido. Use YYYY-MM'},
                status=status.HTTP_400_BAD_REQUEST
            )

def descargar_template_compras(request):
    """
    Vista para descargar el template de importación de compras actualizado
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Template Compras"

        headers = [
            'empresa',
            'proveedor',
            'almacen',
            'fecha',
            'producto',
            'nombre_producto',
            'cantidad',
            'precio_unitario',
            'subtotal',
            'igv_incluido',
            'estado',
            'metodo_pago'
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=template_compras.xlsx'
        wb.save(response)
        return response
    except Exception as e:
        logger.error("Error al generar template de compras: %s", e, exc_info=True)
        return HttpResponse(status=500)

class PagoCompraViewSet(viewsets.ModelViewSet):
    queryset = PagoCompra.objects.all()
    serializer_class = PagoCompraSerializer
    permission_classes = [IsAuthenticated, HasEmpresaPermission]

    def get_queryset(self):
        return self.queryset.filter(compra__empresa=self.request.user.empresa)

    def perform_create(self, serializer):
        compra = serializer.validated_data['compra']
        if compra.empresa != self.request.user.empresa:
            raise PermissionDenied("No tienes permiso para crear pagos para esta compra")
        serializer.save()

        # Actualizar el total pagado y estado de la compra
        compra.pagos_total = compra.pagos.aggregate(total=Sum('monto'))['total'] or 0
        if compra.pagos_total >= compra.total:
            compra.estado = 'pagada'
        compra.save()

    def perform_update(self, serializer):
        compra = serializer.validated_data['compra']
        if compra.empresa != self.request.user.empresa:
            raise PermissionDenied("No tienes permiso para actualizar pagos de esta compra")
        serializer.save()

        # Actualizar el total pagado y estado de la compra
        compra.pagos_total = compra.pagos.aggregate(total=Sum('monto'))['total'] or 0
        if compra.pagos_total >= compra.total:
            compra.estado = 'pagada'
        elif compra.pagos_total < compra.total:
            compra.estado = 'pendiente'
        compra.save()

    def perform_destroy(self, instance):
        compra = instance.compra
        if compra.empresa != self.request.user.empresa:
            raise PermissionDenied("No tienes permiso para eliminar pagos de esta compra")
        instance.delete()

        # Actualizar el total pagado y estado de la compra
        compra.pagos_total = compra.pagos.aggregate(total=Sum('monto'))['total'] or 0
        if compra.pagos_total < compra.total:
            compra.estado = 'pendiente'
        compra.save()
    