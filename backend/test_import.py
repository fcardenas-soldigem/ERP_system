import os
import django
import pandas as pd
from openpyxl import Workbook
from django.core.files.uploadedfile import SimpleUploadedFile

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from apps.inventario.views import importar_productos
from rest_framework.request import Request
from django.contrib.auth import get_user_model
from apps.inventario.models import Categoria, Almacen
from django.http import HttpRequest
from django.middleware.csrf import get_token
from rest_framework.test import APIRequestFactory

def obtener_ids_empresa(user):
    """Obtener IDs de categorías y almacenes de la empresa"""
    categorias = Categoria.objects.filter(empresa=user.empresa)
    almacenes = Almacen.objects.filter(empresa=user.empresa)
    
    print("\nCategorías disponibles:")
    for cat in categorias:
        print(f"- ID: {cat.id}, Nombre: {cat.nombre}")
    
    print("\nAlmacenes disponibles:")
    for alm in almacenes:
        print(f"- ID: {alm.id}, Nombre: {alm.nombre}")
    
    if not categorias or not almacenes:
        return None, None
    
    return categorias.first().id, almacenes.first().id

def crear_template_excel(categoria_id, almacen_id):
    """Crear un archivo Excel de prueba con datos de productos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Definir encabezados
    headers = [
        'SKU', 'nombre', 'descripcion', 'categoria', 'almacen',
        'precio_compra', 'precio_venta', 'stock_total',
        'stock_minimo', 'stock_maximo', 'is_active'
    ]
    
    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Datos de prueba usando los IDs correctos
    test_data = [
        [f'SKU001', 'Producto 1', 'Descripción 1', categoria_id, almacen_id, 100, 150, 50, 10, 100, True],
        [f'SKU002', 'Producto 2', 'Descripción 2', categoria_id, almacen_id, 200, 300, 30, 5, 50, True],
        [f'SKU003', 'Producto 3', 'Descripción 3', categoria_id, almacen_id, 150, 225, 20, 5, 40, True],
    ]

    # Escribir datos
    for row, data in enumerate(test_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Guardar archivo
    filename = 'test_productos.xlsx'
    wb.save(filename)
    return filename

def crear_datos_prueba(user):
    """Crear categoría y almacén de prueba si no existen"""
    # Crear categoría de prueba
    categoria = Categoria.objects.get_or_create(
        nombre='Categoría de Prueba',
        empresa=user.empresa,
        defaults={
            'descripcion': 'Categoría creada para pruebas de importación'
        }
    )[0]

    # Crear almacén de prueba
    almacen = Almacen.objects.get_or_create(
        nombre='Almacén de Prueba',
        empresa=user.empresa,
        defaults={
            'direccion': 'Dirección de prueba'
        }
    )[0]

    return categoria.id, almacen.id

def test_importacion():
    """Probar la importación de productos"""
    try:
        # Usar un usuario existente buscando por email
        User = get_user_model()
        user = User.objects.get(email='fabriziocardenas7@gmail.com')

        print(f"\nUsuario encontrado: {user.email}")
        print(f"Empresa: {user.empresa.nombre if user.empresa else 'Sin empresa'}")

        if not user.empresa:
            print("\nError: El usuario no tiene una empresa asignada.")
            return

        # Crear datos de prueba
        categoria_id, almacen_id = crear_datos_prueba(user)
        print("\nDatos de prueba creados:")
        print(f"- Categoría ID: {categoria_id}")
        print(f"- Almacén ID: {almacen_id}")

        # Mostrar todos los datos disponibles
        obtener_ids_empresa(user)

        # Crear archivo Excel de prueba
        filename = crear_template_excel(categoria_id, almacen_id)

        # Leer el archivo y crear SimpleUploadedFile
        with open(filename, 'rb') as f:
            file_content = f.read()
        
        uploaded_file = SimpleUploadedFile(
            name='test_productos.xlsx',
            content=file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Crear request de prueba usando APIRequestFactory
        factory = APIRequestFactory()
        request = factory.post('/api/inventario/productos/importar/', {'file': uploaded_file})
        request.user = user

        # Ejecutar importación
        response = importar_productos(request)
        
        # Limpiar archivo de prueba
        os.remove(filename)

        # Imprimir resultados
        print("\nResultado de la importación:")
        print(f"Status Code: {response.status_code}")
        print(f"Response Data: {response.data}")

    except User.DoesNotExist:
        print("\nError: Usuario con email 'fabriziocardenas7@gmail.com' no encontrado.")
    except Exception as e:
        print(f"\nError: {str(e)}")

if __name__ == '__main__':
    test_importacion() 